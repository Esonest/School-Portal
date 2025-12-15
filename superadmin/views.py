from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from accounts.decorators import superadmin_required
from accounts.models import User, School, Teacher, SchoolAdmin



# super_admin/views.py
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from students.models import Student
from .forms import SchoolForm, CreateAndAssignAdminForm


from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import Http404
from django.contrib.auth import get_user_model

import json
from django.db.models import Count
from django.utils.timezone import now, timedelta


@login_required
def super_admin_dashboard(request):
    user = request.user
    if not user.is_superuser:
        raise Http404("Not allowed")

    # --- Counts ---
    total_schools = School.objects.count()
    total_school_admins = User.objects.filter(school_admin_profile__isnull=False).count()
    total_teachers = Teacher.objects.count()
    total_students = Student.objects.count()

    # --- Recent entries ---
    recent_schools = School.objects.order_by('-id')[:5]
    recent_teachers = Teacher.objects.select_related('user').order_by('-id')[:5]
    recent_students = Student.objects.select_related('user').order_by('-id')[:5]

    # --- Chart Data ---
    # Example: schools created per day for last 7 days
    today = now().date()
    last_7_days = [today - timedelta(days=i) for i in range(6, -1, -1)]
    school_growth_dates = [d.strftime("%b %d") for d in last_7_days]
    school_growth_counts = []
    for d in last_7_days:
        count = School.objects.filter(created_on__date=d).count()
        school_growth_counts.append(count)

    # Pass chart data as JSON
    school_growth_json = json.dumps(school_growth_counts)

    context = {
        'total_schools': total_schools,
        'total_school_admins': total_school_admins,
        'total_teachers': total_teachers,
        'total_students': total_students,
        'recent_schools': recent_schools,
        'recent_teachers': recent_teachers,
        'recent_students': recent_students,
        'school_growth_json': school_growth_json,
        'school_growth_dates': json.dumps(school_growth_dates),
        'school_count': total_schools,
        'teacher_count': total_teachers,
        'student_count': total_students,
    }

    return render(request, 'superadmin/dashboard.html', context)

@superadmin_required
def school_list(request):
    schools = School.objects.all().order_by('name')
    return render(request, 'superadmin/schools/list.html', {"schools": schools})

@superadmin_required
def school_create(request):
    if request.method == "POST":
        form = SchoolForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "School created successfully.")
            return redirect("superadmin:school_list")
    else:
        form = SchoolForm()

    return render(request, "superadmin/schools/create.html", {"form": form})



@superadmin_required
def school_edit(request, pk):
    school = get_object_or_404(School, pk=pk)

    if request.method == "POST":
        form = SchoolForm(request.POST, request.FILES, instance=school)
        if form.is_valid():
            form.save()
            messages.success(request, "School updated successfully.")
            return redirect("superadmin:school_list")
    else:
        form = SchoolForm(instance=school)

    return render(request, "superadmin/schools/edit.html", {"form": form, "school": school})



@superadmin_required
def school_delete(request, pk):
    school = get_object_or_404(School, pk=pk)

    if request.method == "POST":
        school.delete()
        messages.success(request, "School deleted successfully.")
        return redirect("superadmin:school_list")

    return render(request, "superadmin/schools/delete_confirm.html", {"school": school})



# -----------------------
# SCHOOL ADMIN CRUD
# -----------------------




@superadmin_required
def admin_list(request):
    # Get all users with role='schooladmin'
    admins = User.objects.filter(role='schooladmin').order_by('username')
    return render(request, "superadmin/admins/list.html", {"admins": admins})





@superadmin_required
def admin_create(request):
    schools = School.objects.all()

    if request.method == "POST":
        # Get form data
        name = request.POST.get("name")
        username = request.POST.get("username")
        email = request.POST.get("email")
        password = request.POST.get("password")
        school_id = request.POST.get("school")

        school = None
        if school_id:
            school = get_object_or_404(School, id=school_id)

        # Create the user with role='schooladmin'
        user = User.objects.create_user(
            username=username,
            password=password,
            role='schooladmin',
            school=school
        )

        # Save name and email
        user.first_name = name  # or split into first/last if you want
        user.email = email
        user.save()

        # Optional: create SchoolAdminProfile
        SchoolAdmin.objects.create(
            user=user,
            school=school
        )

        messages.success(request, "School Admin created successfully.")
        return redirect("superadmin:admin_list")


    return render(request, "superadmin/admins/create_admin.html", {"schools": schools})
    




@superadmin_required
def admin_edit(request, pk):
    # Get the user with role='schooladmin'
    admin_user = get_object_or_404(User, pk=pk, role='schooladmin')
    schools = School.objects.all()

    if request.method == "POST":
        form = CreateAndAssignAdminForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            admin_user.username = data['username']
            admin_user.first_name = data['first_name']
            admin_user.last_name = data['last_name']
            admin_user.email = data['email']
            if data['password']:
                admin_user.set_password(data['password'])
            admin_user.school = data.get('school')
            admin_user.save()

            # Update or create SchoolAdminProfile if school is assigned
            if data.get('school'):
                profile, created = SchoolAdmin.objects.get_or_create(user=admin_user)
                profile.school = data['school']
                profile.save()

            messages.success(request, "School Admin updated successfully.")
            return redirect("superadmin:admin_list")
    else:
        # Pre-fill form with existing data
        initial_data = {
            "username": admin_user.username,
            "first_name": admin_user.first_name,
            "last_name": admin_user.last_name,
            "email": admin_user.email,
            "school": admin_user.school
        }
        form = CreateAndAssignAdminForm(initial=initial_data)

    return render(request, "superadmin/admins/edit_admin.html", {"form": form, "admin_user": admin_user, "schools": schools})


@superadmin_required
def admin_delete(request, pk):
    # Get the user with role='schooladmin'
    admin_user = get_object_or_404(User, pk=pk, role='schooladmin')

    if request.method == "POST":
        # Delete the user (and related profiles via cascade if defined)
        admin_user.delete()
        messages.success(request, "School Admin deleted successfully.")
        return redirect("superadmin:admin_list")

    return render(request, "superadmin/admins/delete_admin.html", {"admin_user": admin_user})


@superadmin_required
def teacher_list(request):
    teachers = Teacher.objects.select_related("user", "school")
    return render(request, "superadmin/teachers/list.html", {"teachers": teachers})


# superadmin/views/student_views.py
import csv, io, secrets
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.utils import timezone

from students.models import Student, SchoolClass
from accounts.models import School
from results.models import Score  # may exist
from attendance.models import Attendance  # may exist
from .forms import StudentUserForm, StudentProfileForm, BulkUploadForm

User = get_user_model()

@login_required
def student_list(request):
    qs = Student.objects.select_related('user', 'school', 'school_class').all()

    # Search filter
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            models.Q(user__first_name__icontains=q) |
            models.Q(user__last_name__icontains=q) |
            models.Q(admission_no__icontains=q) |
            models.Q(user__username__icontains=q)
        )

    # filter by class
    class_id = request.GET.get('class_id')
    if class_id:
        qs = qs.filter(school_class_id=class_id)

    # filter by status
    status = request.GET.get('status')
    if status:
        qs = qs.filter(status=status)

    # pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(qs.order_by('-created_at'), 25)
    students = paginator.get_page(page)

    classes = SchoolClass.objects.all()
    return render(request, 'superadmin/students/list.html', {
        'students': students,
        'classes': classes,
        'q': q,
        'class_id': class_id,
        'status': status,
    })

from results.utils import generate_unique_username

@login_required
@transaction.atomic
def create_student(request):
    if request.method == 'POST':
        uform = StudentUserForm(request.POST, request.FILES)
        pform = StudentProfileForm(request.POST, request.FILES)

        if uform.is_valid() and pform.is_valid():
            user = uform.save(commit=False)

            # Generate unique username if blank
            user.username = generate_unique_username(user.username or user.first_name)

            # Set password
            pw = uform.cleaned_data.get('password')
            if pw:
                user.set_password(pw)
                temp_pw = pw
            else:
                temp_pw = secrets.token_urlsafe(8)
                user.set_password(temp_pw)

            user.is_student = True
            user.save()

            # Save student profile
            student = pform.save(commit=False)
            student.user = user
            student.save()

            messages.success(request, f"Student '{user.get_full_name()}' created. Temporary password: {temp_pw}")
            return redirect('superadmin:student_list')
        else:
            # Show form errors
            messages.error(request, f"User form errors: {uform.errors}, Profile form errors: {pform.errors}")
    else:
        uform = StudentUserForm()
        pform = StudentProfileForm(initial={'session': f"{timezone.now().year}/{timezone.now().year+1}"})

    return render(request, 'superadmin/students/create.html', {
        'form': uform,
        'profile_form': pform
    })


@login_required
@transaction.atomic
def edit_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    user = student.user
    if request.method == 'POST':
        uform = StudentUserForm(request.POST, request.FILES, instance=user)
        pform = StudentProfileForm(request.POST, request.FILES, instance=student)
        if uform.is_valid() and pform.is_valid():
            u = uform.save(commit=False)
            pw = uform.cleaned_data.get('password')
            if pw:
                u.set_password(pw)
            u.save()
            pform.save()
            messages.success(request, "Student updated")
            return redirect('superadmin:student_list')
    else:
        uform = StudentUserForm(instance=user)
        pform = StudentProfileForm(instance=student)
    return render(request, 'superadmin/students/edit.html', {'form': uform, 'profile_form': pform, 'student': student})

@login_required
@transaction.atomic
def bulk_upload_students(request):
    if request.method == 'POST':
        form = BulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            # parse csv
            f = form.cleaned_data['file']
            decoded = f.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(decoded))
            created = 0; errors = []
            for idx, row in enumerate(reader, start=1):
                try:
                    first_name = row.get('first_name') or row.get('fname') or ''
                    last_name = row.get('last_name') or row.get('lname') or ''
                    email = row.get('email') or ''
                    school_id = row.get('school_id') or row.get('school') or ''
                    class_id = row.get('class_id') or row.get('class') or ''
                    dob = row.get('dob') or None
                    gender = row.get('gender') or ''
                    session = row.get('session') or ''
                    term = row.get('term') or ''
                    # get school/class
                    school = School.objects.filter(id=school_id).first() if school_id else None
                    school_class = SchoolClass.objects.filter(id=class_id).first() if class_id else None
                    # create user
                    username = (first_name[:1] + last_name).lower() + secrets.token_hex(2)
                    user = User.objects.create(username=username, first_name=first_name, last_name=last_name, email=email)
                    pw = secrets.token_urlsafe(8)
                    user.set_password(pw); user.is_student = True; user.save()
                    student = Student.objects.create(user=user, school=school or None, school_class=school_class or None, dob=dob or None, gender=gender, session=session, term=term)
                    created += 1
                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")
            messages.success(request, f"Bulk upload finished. Created: {created}. Errors: {len(errors)}")
            if errors:
                messages.error(request, "Sample errors: " + (errors[0] if errors else ""))
            return redirect('superadmin:student_list')
    else:
        form = BulkUploadForm()
    return render(request, 'superadmin/students/bulk_upload.html', {'form': form})

@login_required
def student_profile(request, student_id):
    student = get_object_or_404(Student.objects.select_related('user','school','school_class'), id=student_id)

    # Attendance stats (fallbacks if Attendance absent)
    attendance_stats = {'present': 0, 'absent': 0, 'percentage': None}
    try:
        total = Attendance.objects.filter(student=student).count()
        present = Attendance.objects.filter(student=student, status='present').count()
        absent = total - present
        attendance_stats['present'] = present
        attendance_stats['absent'] = absent
        attendance_stats['percentage'] = round((present/total)*100,2) if total>0 else None
    except Exception:
        # attendance app not installed or model different
        attendance_stats = None

    # Performance summary from Score model
    perf = {}
    try:
        scores = Score.objects.filter(student=student)
        if scores.exists():
            # overall average
            avg = scores.annotate(total=models.ExpressionWrapper(models.F('ca')+models.F('exam'), output_field=models.FloatField())).aggregate(avg_total=models.Avg('total'))['avg_total']
            perf['average'] = round(avg or 0,2)
            # subjects breakdown
            subjects = []
            for s in scores.select_related('subject'):
                subjects.append({'subject': s.subject.name, 'total': round((s.ca or 0)+(s.exam or 0),2), 'grade': s.grade})
            perf['subjects'] = subjects
        else:
            perf = None
    except Exception:
        perf = None

    return render(request, 'superadmin/students/profile.html', {
        'student': student,
        'attendance_stats': attendance_stats,
        'performance': perf
    })

@login_required
def delete_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)

    if request.method == "POST":
        user = student.user
        student.delete()
        user.delete()

        messages.success(request, "Student deleted successfully.")
        return redirect("superadmin:student_list")

    return render(request, "superadmin/students/delete.html", {
        "student": student
    })



# superadmin/views/teacher_views.py

from django.shortcuts import render, redirect, get_object_or_404
from accounts.models import Teacher
from .forms import TeacherForm, TeacherImportForm

# 🟢 Create teacher
def teacher_create(request):
    form = TeacherForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        teacher = form.save(commit=False)
        teacher.save()
        form.save_m2m()

        # 🔥 Ensure initial mapping exists
        from results.utils import sync_class_subject_teacher
        sync_class_subject_teacher(teacher)

        return redirect("superadmin:teacher_list")

    return render(
        request,
        "superadmin/teachers/teacher_form.html",
        {"form": form}
    )




# 🟢 List all teachers
def teacher_list(request):
    teachers = Teacher.objects.select_related("school", "user").prefetch_related("subjects", "classes")
    return render(request, "superadmin/teachers/teacher_list.html", {"teachers": teachers})

# 🟢 Edit teacher
def teacher_edit(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    form = TeacherForm(request.POST or None, instance=teacher)

    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("superadmin:teacher_list")

    return render(request, "superadmin/teachers/teacher_form.html", {"form": form})

# 🔴 Delete teacher
@login_required
def teacher_delete(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)

    if request.method == "POST":
        user = teacher.user
        teacher.delete()
        user.delete()

        messages.success(request, "Teacher deleted successfully.")
        return redirect("superadmin:teacher_list")

    return render(request, "superadmin/teachers/delete_teacher.html", {
        "teacher": teacher
    })


from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import Q
from results.models import Subject

def teacher_list_advanced(request):
    teachers = Teacher.objects.select_related("user", "school")\
        .prefetch_related("subjects", "classes")

    # --- SEARCH ---
    search = request.GET.get("search")
    if search:
        teachers = teachers.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search) |
            Q(subjects__name__icontains=search) |
            Q(classes__name__icontains=search)
        ).distinct()

    # --- FILTER BY SCHOOL ---
    school_id = request.GET.get("school")
    if school_id:
        teachers = teachers.filter(school_id=school_id)

    # --- FILTER BY CLASS ---
    class_id = request.GET.get("class")
    if class_id:
        teachers = teachers.filter(classes__id=class_id)

    # --- FILTER BY SUBJECT ---
    subject_id = request.GET.get("subject")
    if subject_id:
        teachers = teachers.filter(subjects__id=subject_id)

    # --- PAGINATION ---
    paginator = Paginator(teachers, 10)
    page_number = request.GET.get("page")
    teacher_page = paginator.get_page(page_number)

    context = {
        "teachers": teacher_page,
        "schools": School.objects.all(),
        "classes": SchoolClass.objects.all(),
        "subjects": Subject.objects.all(),
        "search": search or "",
        "selected_school": school_id,
        "selected_class": class_id,
        "selected_subject": subject_id,
    }
    return render(request, "superadmin/teachers/teacher_list_advanced.html", context)


import openpyxl
from django.contrib import messages
from django.shortcuts import redirect

def teacher_import(request):
    if request.method == "POST":
        upload = request.FILES["file"]
        wb = openpyxl.load_workbook(upload)
        ws = wb.active

        count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            full_name, email, school_name = row[:3]

            if not email:
                continue

            # Get or create user
            user, created = User.objects.get_or_create(
                email=email,
                defaults={"username": email, "first_name": full_name},
            )

            # Get school
            school = School.objects.filter(name__iexact=school_name).first()
            if not school:
                continue

            # Create teacher
            Teacher.objects.get_or_create(user=user, school=school)
            count += 1

        messages.success(request, f"{count} teachers imported successfully!")
        return redirect("superadmin:teacher_list_advanced")

    form = TeacherImportForm()
    return render(request, "superadmin/teachers/teacher_import.html", {"form": form})

import openpyxl
from django.http import HttpResponse

def export_teachers_excel(request):
    teachers = Teacher.objects.select_related("user", "school").prefetch_related("subjects", "classes")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Teachers"

    ws.append(["Full Name", "Email", "School", "Subjects", "Classes"])

    for t in teachers:
        ws.append([
            t.user.get_full_name(),
            t.user.email,
            t.school.name,
            ", ".join(s.name for s in t.subjects.all()),
            ", ".join(c.name for c in t.classes.all()),
        ])

    response = HttpResponse(
        content=openpyxl.writer.excel.save_virtual_workbook(wb),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=teachers.xlsx"
    return response

from .forms import SchoolClassForm, SubjectForm


@login_required
def class_list(request):
    if not request.user.is_superadmin:
        raise PermissionDenied

    classes = SchoolClass.objects.select_related("school").order_by("name")
    return render(request, "superadmin/classes/class_list.html", {"classes": classes})


@login_required
def class_create(request):
    if not request.user.is_superadmin:
        raise PermissionDenied

    if request.method == "POST":
        form = SchoolClassForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Class created successfully.")
            return redirect("superadmin:class_list")
    else:
        form = SchoolClassForm()

    return render(request, "superadmin/classes/class_form.html", {"form": form})


@login_required
def class_edit(request, class_id):
    if not request.user.is_superadmin:
        raise PermissionDenied

    school_class = get_object_or_404(SchoolClass, id=class_id)

    if request.method == "POST":
        form = SchoolClassForm(request.POST, instance=school_class)
        if form.is_valid():
            form.save()
            messages.success(request, "Class updated successfully.")
            return redirect("superadmin:class_list")
    else:
        form = SchoolClassForm(instance=school_class)

    return render(request, "superadmin/classes/class_form.html", {"form": form})


@login_required
def class_delete(request, class_id):
    if not request.user.is_superadmin:
        raise PermissionDenied

    school_class = get_object_or_404(SchoolClass, id=class_id)
    school_class.delete()
    messages.success(request, "Class deleted successfully.")

    return redirect("superadmin:class_list")

@login_required
def subject_list(request):
    if not request.user.is_superadmin:
        raise PermissionDenied

    subjects = Subject.objects.select_related("school").order_by("name")
    return render(request, "superadmin/subjects/subject_list.html", {"subjects": subjects})


@login_required
def subject_create(request):
    if not request.user.is_superadmin:
        raise PermissionDenied

    if request.method == "POST":
        form = SubjectForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Subject created successfully.")
            return redirect("superadmin:subject_list")
    else:
        form = SubjectForm()

    return render(request, "superadmin/subjects/subject_form.html", {"form": form})


@login_required
def subject_edit(request, subject_id):
    if not request.user.is_superadmin:
        raise PermissionDenied

    subject = get_object_or_404(Subject, id=subject_id)

    if request.method == "POST":
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            form.save()
            messages.success(request, "Subject updated successfully.")
            return redirect("superadmin:subject_list")
    else:
        form = SubjectForm(instance=subject)

    return render(request, "superadmin/subjects/subject_form.html", {"form": form})


@login_required
def subject_delete(request, subject_id):
    if not request.user.is_superadmin:
        raise PermissionDenied

    subject = get_object_or_404(Subject, id=subject_id)
    subject.delete()
    messages.success(request, "Subject deleted successfully.")

    return redirect("superadmin:subject_list")


from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.db.models import Q
from django.contrib import messages
from django.contrib.auth import get_user_model

from .forms import UserCreateForm, UserUpdateForm

User = get_user_model()


def user_list(request):
    search = request.GET.get("search", "")
    role = request.GET.get("role", "")
    status = request.GET.get("status", "")

    users = User.objects.all().order_by("-id")

    # SEARCH
    if search:
        users = users.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(username__icontains=search) |
            Q(email__icontains=search)
        )

    # FILTER BY ROLE
    if role:
        users = users.filter(role=role)

    # FILTER ACTIVE/INACTIVE
    if status == "active":
        users = users.filter(is_active=True)
    elif status == "inactive":
        users = users.filter(is_active=False)

    # PAGINATION
    paginator = Paginator(users, 10)
    page = request.GET.get("page")
    users_page = paginator.get_page(page)

    return render(request, "superadmin/users/user_list.html", {
        "users": users_page,
        "search": search,
        "role": role,
        "status": status,
    })


from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import Http404

# accounts/views.py
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .forms import UserCreateForm

@login_required
def user_create(request):
    if request.method == "POST":
        form = UserCreateForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f"User {user.username} created successfully!")
            return redirect('superadmin:user_list')  # replace with your users list URL
    else:
        form = UserCreateForm()

    return render(request, "superadmin/users/user_form.html", {"form": form})



def user_update(request, user_id):
    user = get_object_or_404(User, id=user_id)
    form = UserUpdateForm(request.POST or None, instance=user)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "User updated successfully.")
        return redirect("superadmin:user_list")

    return render(request, "superadmin/users/user_form.html", {"form": form})


def user_delete(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.delete()
    messages.success(request, "User deleted successfully.")
    return redirect("superadmin:user_list")


def user_detail(request, user_id):
    user = get_object_or_404(User, id=user_id)

    # LINKED PROFILES
    student = getattr(user, "student_profile", None)
    teacher = getattr(user, "teacher_profile", None)
    admin = getattr(user, "admin_profile", None)

    return render(request, "superadmin/users/user_detail.html", {
        "user_obj": user,
        "student": student,
        "teacher": teacher,
        "admin": admin,
    })




from django.contrib.auth.decorators import user_passes_test
from .models import SchoolPortalSetting
from .forms import SchoolPortalSettingForm


def is_superadmin(user):
    return user.is_superuser


@user_passes_test(is_superadmin)
def school_portal_setting_list(request):
    schools = School.objects.select_related("portal_settings")

    return render(request, "superadmin/portal_settings/portal_setting_list.html", {
        "schools": schools
    })


@user_passes_test(is_superadmin)
def school_portal_setting_update(request, school_id):
    school = get_object_or_404(School, id=school_id)
    settings, _ = SchoolPortalSetting.objects.get_or_create(school=school)

    if request.method == "POST":
        form = SchoolPortalSettingForm(request.POST, instance=settings)
        if form.is_valid():
            form.save()
            messages.success(request, "Portal settings updated successfully.")
            return redirect("superadmin:superadmin_portal_list")
    else:
        form = SchoolPortalSettingForm(instance=settings)

    return render(request, "superadmin/portal_settings/portal_setting_form.html", {
        "school": school,
        "form": form
    })
