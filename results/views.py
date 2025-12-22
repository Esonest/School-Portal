from django.shortcuts import render, get_object_or_404
from .models import Student, Score, Psychomotor, Affective, School, Subject, ClassSubjectTeacher
import random
import qrcode
from io import BytesIO
from .utils import portal_required, save_qr_to_student, generate_verification_qr
from django.templatetags.static import static


# Result portal for a student

def student_result(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    scores = Score.objects.filter(student=student)
    total_marks = sum([s.total for s in scores])
    average_score = total_marks / scores.count() if scores.exists() else 0
    best_subject = max(scores, key=lambda x: x.total).subject.name if scores.exists() else None
    least_subject = min(scores, key=lambda x: x.total).subject.name if scores.exists() else None

    # Random comments
    principal_comments = [
        f"{student.user.first_name}, keep up the great work!",
        f"{student.user.first_name}, strive for excellence in all subjects.",
        f"{student.user.first_name}, well done this term!"
    ]
    teacher_comments = [
        f"{student.user.first_name}, you have improved a lot this term.",
        f"{student.user.first_name}, keep focusing on your weak areas.",
        f"{student.user.first_name}, excellent effort in class."
    ]

    # QR Code generation
    qr = qrcode.QRCode(box_size=2, border=2)
    qr.add_data(f"Student ID: {student.id} - Total: {total_marks}")
    qr.make(fit=True)
    img = qr.make_image()
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    qr_code_base64 = base64.b64encode(buffer.getvalue()).decode()

    context = {
        'student': student,
        'scores': scores,
        'total_marks': total_marks,
        'average_score': round(average_score,2),
        'best_subject': best_subject,
        'least_subject': least_subject,
        'principal_comment': random.choice(principal_comments),
        'teacher_comment': random.choice(teacher_comments),
        'qr_code': qr_code_base64,
    }
    return render(request, 'results/student_result.html', context)

def student_result(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    scores = Score.objects.filter(student=student)
    total_marks = sum([s.total for s in scores])
    average_score = total_marks / scores.count() if scores.exists() else 0

    # Determine overall grade
    if average_score >= 70:
        overall_grade = 'A'
    elif average_score >= 60:
        overall_grade = 'B'
    elif average_score >= 50:
        overall_grade = 'C'
    elif average_score >= 45:
        overall_grade = 'D'
    else:
        overall_grade = 'F'

    best_subject = max(scores, key=lambda x: x.total).subject.name if scores.exists() else None
    least_subject = min(scores, key=lambda x: x.total).subject.name if scores.exists() else None

    # Comments based on grade
    principal_comments_dict = {
        'A': [f"{student.user.first_name}, excellent performance! Keep it up!",
              f"{student.user.first_name}, outstanding achievement this term!"],
        'B': [f"{student.user.first_name}, good job! Aim higher next term.",
              f"{student.user.first_name}, solid effort! Maintain consistency."],
        'C': [f"{student.user.first_name}, satisfactory performance, more effort is needed.",
              f"{student.user.first_name}, focus on weak subjects next term."],
        'D': [f"{student.user.first_name}, needs improvement. Let’s work harder!",
              f"{student.user.first_name}, concentrate and improve performance."],
        'F': [f"{student.user.first_name}, poor performance. Immediate improvement required!",
              f"{student.user.first_name}, please seek help and work harder!"]
    }

    teacher_comments_dict = {
        'A': [f"{student.user.first_name}, excellent class participation and results.",
              f"{student.user.first_name}, a role model in class."],
        'B': [f"{student.user.first_name}, good effort, can still improve.",
              f"{student.user.first_name}, focus on challenging topics."],
        'C': [f"{student.user.first_name}, needs to put in more effort.",
              f"{student.user.first_name}, practice consistently to improve."],
        'D': [f"{student.user.first_name}, slow progress, extra support needed.",
              f"{student.user.first_name}, must improve dedication."],
        'F': [f"{student.user.first_name}, poor engagement. Immediate attention required.",
              f"{student.user.first_name}, remedial support recommended."]
    }

    import random
    principal_comment = random.choice(principal_comments_dict[overall_grade])
    teacher_comment = random.choice(teacher_comments_dict[overall_grade])

    # QR Code generation
    import qrcode
    from io import BytesIO
    import base64
    qr = qrcode.QRCode(box_size=2, border=2)
    qr.add_data(f"Student ID: {student.id} - Total: {total_marks}")
    qr.make(fit=True)
    img = qr.make_image()
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    qr_code_base64 = base64.b64encode(buffer.getvalue()).decode()

    context = {
        'student': student,
        'scores': scores,
        'total_marks': total_marks,
        'average_score': round(average_score,2),
        'best_subject': best_subject,
        'least_subject': least_subject,
        'principal_comment': principal_comment,
        'teacher_comment': teacher_comment,
        'qr_code': qr_code_base64,
    }
    return render(request, 'results/student_result.html', context)


from django.forms import modelformset_factory
from django.shortcuts import redirect



from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.forms import modelformset_factory
from django.apps import apps
from django.contrib.auth.decorators import login_required

Score = apps.get_model('results', 'Score')
Student = apps.get_model('students', 'Student')
SchoolClass = apps.get_model('students', 'SchoolClass')
School = apps.get_model('accounts', 'School')
# Subject may be optional in your project
try:
    Subject = apps.get_model('results', 'Subject')
except LookupError:
    Subject = None

from .forms import ScoreBulkForm



from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib.auth.decorators import login_required
from .utils import SESSION_LIST  # import the session list


@login_required
def bulk_score_entry(request, school_id):
    teacher = getattr(request.user, 'teacher_profile', None)
    school = get_object_or_404(School, id=school_id)

    setting, _ = SystemSetting.objects.get_or_create(
        id=1,
        defaults={"current_session": SESSION_LIST[0], "current_term": "1"}
    )
    current_session = setting.current_session
    current_term = setting.current_term

    selected_session = request.GET.get("session") or request.POST.get("session") or current_session
    selected_term = request.GET.get("term") or request.POST.get("term") or current_term

    # Classes
    try:
        classes_qs = SchoolClass.objects.filter(teacher=teacher, school=school)
    except:
        try:
            classes_qs = teacher.classes.filter(school=school)
        except:
            classes_qs = SchoolClass.objects.filter(school=school)

    # Subjects
    if teacher:
        try:
            subjects_qs = Subject.objects.filter(teacher=teacher, school=school)
        except:
            subjects_qs = Subject.objects.filter(school=school)
    else:
        subjects_qs = Subject.objects.filter(school=school)

    selected_class_id = request.GET.get('class_id') or request.POST.get('class_id')
    selected_subject_ids = request.GET.getlist('subject_id') or request.POST.getlist('subject_id')
    page = request.GET.get('page') or request.POST.get('page') or 1

    selected_class = None
    if selected_class_id:
        try:
            selected_class = classes_qs.get(id=int(selected_class_id))
        except:
            selected_class = None

    if not selected_class and classes_qs.count() == 1:
        selected_class = classes_qs.first()

    selected_subjects = Subject.objects.filter(id__in=selected_subject_ids) if selected_subject_ids else []
    selected_subject_ids = [int(sid) for sid in selected_subject_ids]

    # Students
    students_qs = Student.objects.filter(school_class=selected_class).order_by(
        'user__first_name', 'user__last_name'
    )
    paginator = Paginator(students_qs, 20)
    try:
        students_page = paginator.page(page)
    except PageNotAnInteger:
        students_page = paginator.page(1)
    except EmptyPage:
        students_page = paginator.page(paginator.num_pages)

    page_students = list(students_page.object_list)
    page_student_ids = [s.id for s in page_students]

    # Ensure scores exist
    if request.method == 'GET' and selected_subject_ids:
        for st in page_students:
            for sid in selected_subject_ids:
                subj = Subject.objects.filter(id=sid).first()
                if subj:
                    Score.objects.get_or_create(
                        student=st,
                        subject=subj,
                        session=selected_session,
                        term=selected_term,
                        school=school,
                        defaults={'ca': 0, 'exam': 0}
                    )

    scores = Score.objects.filter(
        student__id__in=page_student_ids,
        subject__id__in=selected_subject_ids,
        session=selected_session,
        term=selected_term,
        school=school
    )
    scores_map = {}
    for sc in scores:
        scores_map.setdefault(sc.student_id, {})[sc.subject_id] = sc

    # Read class score settings
    is_ca_enabled = True
    ca_max = 40
    exam_max = 60
    mode = 'ca_exam'

    if selected_class:
        try:
            setting = selected_class.score_setting
            is_ca_enabled = setting.uses_ca
            ca_max = setting.ca_max
            exam_max = setting.exam_max
            mode = 'exam_only' if not is_ca_enabled else 'ca_exam'
        except ClassScoreSetting.DoesNotExist:
            pass

    # POST: save scores
    if request.method == 'POST' and 'delete_student' not in request.POST:
        errors = []

        for st in page_students:
            for sid in selected_subject_ids:
                # --- SAFE float conversion ---
                exam_val_str = request.POST.get(f"exam_{st.id}_{sid}", "")
                exam_val = float(exam_val_str) if exam_val_str.strip() else 0

                ca_val = 0
                if is_ca_enabled:
                    ca_val_str = request.POST.get(f"ca_{st.id}_{sid}", "")
                    ca_val = float(ca_val_str) if ca_val_str.strip() else 0
                # ---------------------------

                if not (0 <= exam_val <= exam_max):
                    errors.append(f"Exam must be 0–{exam_max} for {st.full_name()}")
                    continue
                if is_ca_enabled and not (0 <= ca_val <= ca_max):
                    errors.append(f"CA must be 0–{ca_max} for {st.full_name()}")
                    continue

                subj = Subject.objects.filter(id=sid).first()
                if not subj:
                    continue

                score_obj, created = Score.objects.get_or_create(
                    student=st,
                    subject=subj,
                    session=selected_session,
                    term=selected_term,
                    school=school,
                    defaults={'ca': ca_val, 'exam': exam_val}
                )
                if not created:
                    score_obj.ca = ca_val
                    score_obj.exam = exam_val
                    score_obj.save()

        # Show messages
        if errors:
            for err in errors:
                messages.error(request, err)
        else:
            messages.success(request, "Scores saved successfully.")

        params = [f'class_id={selected_class.id}']
        for sid in selected_subject_ids:
            params.append(f'subject_id={sid}')
        params.append(f'term={selected_term}')
        params.append(f'session={selected_session}')
        params.append(f'page={students_page.number}')
        return redirect(f"{request.path}?{'&'.join(params)}")

    template_name = 'results/bulk_score_entry.html'

    return render(request, template_name, {
        'school': school,
        'classes': classes_qs,
        'subjects': subjects_qs,
        'selected_class': selected_class,
        'selected_subject_ids': selected_subject_ids,
        'selected_subjects': selected_subjects,
        'students_page': students_page,
        'scores_map': scores_map,
        'paginator': paginator,
        'page_obj': students_page,
        'selected_term': selected_term,
        'term_choices': {'1': 'Term 1', '2': 'Term 2', '3': 'Term 3'},
        'selected_session': selected_session,
        'session_choices': SESSION_LIST,
        'is_ca_enabled': is_ca_enabled,
        'ca_max': ca_max,
        'exam_max': exam_max,
        'mode': mode,
    })








from .models import ClassScoreSetting

@login_required
def class_score_settings(request, school_id):
    school = get_object_or_404(School, id=school_id)
    classes = SchoolClass.objects.filter(school=school).order_by('name')

    # Ensure settings exist for every class
    for cls in classes:
        ClassScoreSetting.objects.get_or_create(
            school_class=cls,
            defaults={
                "ca_max": 0,    # allow zero
                "exam_max": 100 # default dynamic total
            }
        )

    if request.method == "POST":
        for cls in classes:
            setting = ClassScoreSetting.objects.get(school_class=cls)

            # Get values from input – allow zero values
            ca_max = request.POST.get(f"ca_max_{cls.id}", "0")
            exam_max = request.POST.get(f"exam_max_{cls.id}", "0")

            # Convert to int safely
            try:
                ca_max = int(ca_max)
            except:
                ca_max = 0

            try:
                exam_max = int(exam_max)
            except:
                exam_max = 0

            # Save without requiring positive numbers
            setting.ca_max = max(0, ca_max)
            setting.exam_max = max(0, exam_max)
            setting.save()

        messages.success(request, "Score settings updated.")
        return redirect(request.path)

    class_score_settings = ClassScoreSetting.objects.filter(
        school_class__school=school
    ).select_related("school_class")

    return render(request, "results/class_score_settings.html", {
        "school": school,
        "classes": classes,
        "class_score_settings": class_score_settings,
    })






from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_protect
from .models import Score

@csrf_protect
@require_POST
def delete_score(request, student_id, subject_id):
    try:
        deleted, _ = Score.objects.filter(student_id=student_id, subject_id=subject_id).delete()
        if deleted:
            return JsonResponse({'status': 'success'})
        return JsonResponse({'status': 'error', 'message': 'Score not found'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)







from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.forms import modelformset_factory
from django.apps import apps
from django.contrib.auth.decorators import login_required

Psychomotor = apps.get_model('results', 'Psychomotor')
Affective = apps.get_model('results', 'Affective')
Student = apps.get_model('students', 'Student')
SchoolClass = apps.get_model('students', 'SchoolClass')
School = apps.get_model('accounts', 'School')

from .forms import PsychomotorForm, AffectiveForm
from accounts.models import SystemSetting
from .utils import SESSION_LIST


@login_required
def bulk_psycho_affective(request, school_id):

    # ----------------------------
    # SAFE SESSION/TERM HANDLING
    # ----------------------------
    setting = SystemSetting.objects.first()
    if setting:
        current_session = setting.current_session
        current_term = setting.current_term
    else:
        current_session = "2024/2025"
        current_term = "1"

    # ----------------------------
    # SESSION + TERM CAPTURE
    # ----------------------------
    session_choices = SESSION_LIST
    selected_session = (
        request.GET.get("session")
        or request.POST.get("session")
        or current_session
    )

    term_choices = {
        "1": "Term 1",
        "2": "Term 2",
        "3": "Term 3",
    }
    selected_term = (
        request.GET.get("term")
        or request.POST.get("term")
        or current_term
    )

    # ----------------------------
    # USER / SCHOOL CONTEXT
    # ----------------------------
    teacher = getattr(request.user, 'teacher_profile', None)
    school = get_object_or_404(School, id=school_id)

    # ----------------------------
    # CLASSES AVAILABLE TO TEACHER
    # ----------------------------
    if teacher:
        try:
            classes_qs = SchoolClass.objects.filter(teacher=teacher, school=school)
        except:
            try:
                classes_qs = teacher.classes.filter(school=school)
            except:
                classes_qs = SchoolClass.objects.filter(school=school)
    else:
        classes_qs = SchoolClass.objects.filter(school=school)

    selected_class_id = request.GET.get('class_id') or request.POST.get('class_id')
    page_num = request.GET.get('page') or request.POST.get('page') or 1

    selected_class = None
    if selected_class_id:
        try:
            selected_class = classes_qs.get(id=int(selected_class_id))
        except:
            selected_class = None

    if not selected_class and classes_qs.count() == 1:
        selected_class = classes_qs.first()

    if not selected_class:
        return render(request, 'results/bulk_psycho_affective_select.html', {
            'classes': classes_qs,
            'school': school,
            'selected_class': None,
            'session_choices': session_choices,
            'selected_session': selected_session,
            'term_choices': term_choices,
            'selected_term': selected_term,
        })

    # ----------------------------
    # STUDENTS IN CLASS + PAGINATION
    # ----------------------------
    students_qs = Student.objects.filter(
        school_class=selected_class
    ).order_by('user__first_name', 'user__last_name')

    paginator = Paginator(students_qs, 20)

    try:
        students_page = paginator.page(page_num)
    except PageNotAnInteger:
        students_page = paginator.page(1)
    except EmptyPage:
        students_page = paginator.page(paginator.num_pages)

    page_students = list(students_page.object_list)

    # ----------------------------
    # ENSURE PSYCHOMOTOR + AFFECTIVE EXIST
    # ----------------------------
    for st in page_students:
        Psychomotor.objects.get_or_create(
            student=st,
            session=selected_session,
            term=selected_term,
            defaults={
                'neatness': 1,
                'agility': 1,
                'creativity': 1,
                'sports': 1,
                'handwriting': 1,
                'school': school,
            }
        )

        Affective.objects.get_or_create(
            student=st,
            session=selected_session,
            term=selected_term,
            defaults={
                'punctuality': 1,
                'cooperation': 1,
                'behavior': 1,
                'attentiveness': 1,
                'perseverance': 1,
                'school': school,
            }
        )

    # ----------------------------
    # PRELOAD INSTANCES FOR DISPLAY
    # ----------------------------
    psy_instances = [
        Psychomotor.objects.filter(
            student=st, session=selected_session, term=selected_term
        ).first()
        for st in page_students
    ]

    aff_instances = [
        Affective.objects.filter(
            student=st, session=selected_session, term=selected_term
        ).first()
        for st in page_students
    ]

    # ----------------------------
    # BUILD FORM ROWS
    # ----------------------------
    rows = []
    for idx, student in enumerate(page_students):
        psy_inst = psy_instances[idx]
        aff_inst = aff_instances[idx]

        psy_prefix = f'psy-{idx}'
        aff_prefix = f'aff-{idx}'

        if request.method == 'POST':
            psy_form = PsychomotorForm(request.POST, prefix=psy_prefix, instance=psy_inst)
            aff_form = AffectiveForm(request.POST, prefix=aff_prefix, instance=aff_inst)
        else:
            psy_form = PsychomotorForm(prefix=psy_prefix, instance=psy_inst)
            aff_form = AffectiveForm(prefix=aff_prefix, instance=aff_inst)

        rows.append({
            'student': student,
            'psy_form': psy_form,
            'aff_form': aff_form,
            'psy_prefix': psy_prefix,
            'aff_prefix': aff_prefix,
            'index': idx
        })

    # ----------------------------
    # POST HANDLING
    # ----------------------------
    if request.method == 'POST':
        all_valid = all(r['psy_form'].is_valid() and r['aff_form'].is_valid() for r in rows)

        if all_valid:
            for r in rows:
                psy_obj = r['psy_form'].save(commit=False)
                aff_obj = r['aff_form'].save(commit=False)

                psy_obj.school = school
                aff_obj.school = school

                psy_obj.recorded_by = request.user
                aff_obj.recorded_by = request.user

                psy_obj.save()
                aff_obj.save()

            messages.success(request, "Progress saved successfully. You may continue later.")

            params = [
                f'class_id={selected_class.id}',
                f'page={students_page.number}',
                f'term={selected_term}',
                f'session={selected_session}',
            ]
            return redirect(f"{request.path}?{'&'.join(params)}")

        else:
            messages.error(request, "There were validation errors. Please correct and save again.")

    # ----------------------------
    # RENDER PAGE
    # ----------------------------
    context = {
        'school': school,
        'classes': classes_qs,
        'selected_class': selected_class,
        'students_page': students_page,
        'rows': rows,
        'paginator': paginator,
        'page_obj': students_page,

        'session_choices': session_choices,
        'selected_session': selected_session,

        'term_choices': term_choices,
        'selected_term': selected_term,
    }

    return render(request, 'results/bulk_psycho_affective.html', context)








from django.http import HttpResponse
from .utils import render_to_pdf


def generate_termly_report(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    scores = Score.objects.filter(student=student)
    psycho = Psychomotor.objects.get(student=student)
    affective = Affective.objects.get(student=student)
    total_marks = sum([s.total for s in scores])
    average_score = total_marks / scores.count() if scores.exists() else 0

    # Overall grade
    if average_score >= 70: grade = 'A'
    elif average_score >= 60: grade = 'B'
    elif average_score >= 50: grade = 'C'
    elif average_score >= 45: grade = 'D'
    else: grade = 'F'

    # Comments based on grade
    import random
    principal_comments_dict = {
        'A': [f"{student.user.first_name}, excellent performance! Keep it up!"],
        'B': [f"{student.user.first_name}, good job! Aim higher next term."],
        'C': [f"{student.user.first_name}, satisfactory performance, more effort is needed."],
        'D': [f"{student.user.first_name}, needs improvement. Let's work harder!"],
        'F': [f"{student.user.first_name}, poor performance. Immediate improvement required!"]
    }
    teacher_comments_dict = {
        'A': [f"{student.user.first_name}, excellent class participation and results."],
        'B': [f"{student.user.first_name}, good effort, can still improve."],
        'C': [f"{student.user.first_name}, needs to put in more effort."],
        'D': [f"{student.user.first_name}, slow progress, extra support needed."],
        'F': [f"{student.user.first_name}, poor engagement. Immediate attention required."]
    }

    # QR Code generation
    import qrcode
    import base64
    from io import BytesIO
    qr = qrcode.QRCode(box_size=2, border=2)
    qr.add_data(f"Student ID: {student.id} - Total: {total_marks}")
    qr.make(fit=True)
    img = qr.make_image()
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    qr_code_base64 = base64.b64encode(buffer.getvalue()).decode()

    context = {
        'student': student,
        'scores': scores,
        'total_marks': total_marks,
        'average_score': round(average_score,2),
        'best_subject': max(scores, key=lambda x: x.total).subject.name if scores.exists() else None,
        'least_subject': min(scores, key=lambda x: x.total).subject.name if scores.exists() else None,
        'principal_comment': random.choice(principal_comments_dict[grade]),
        'teacher_comment': random.choice(teacher_comments_dict[grade]),
        'psycho': psycho,
        'affective': affective,
        'qr_code': qr_code_base64
    }

    pdf = render_to_pdf('results/termly_report.html', context)
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = f"{student.user.get_full_name()}_Termly_Report.pdf"
        content = f"inline; filename={filename}"
        response['Content-Disposition'] = content
        return response
    return HttpResponse("Error generating PDF")


import pandas as pd
from django.shortcuts import render, redirect
from .models import Student, Score, Subject
from accounts.models import School
from django.contrib import messages
from django.contrib.auth.models import User

# Bulk Student Upload

def bulk_student_upload(request, school_id):
    school = School.objects.get(id=school_id)
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        df = pd.read_csv(file)
        for _, row in df.iterrows():
            user, created = User.objects.get_or_create(username=row['username'], defaults={
                'first_name': row['first_name'],
                'last_name': row['last_name'],
                'email': row.get('email', '')
            })
            Student.objects.get_or_create(
                user=user,
                school=school,
                gender=row['gender'],
                student_class=row['student_class']
            )
        messages.success(request, "Students uploaded successfully")
        return redirect('school_dashboard', school_id=school.id)
    return render(request, 'results/bulk_student_upload.html', {'school': school})

# Bulk Score Upload

def bulk_score_upload(request, school_id):
    school = School.objects.get(id=school_id)
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        df = pd.read_csv(file)
        for _, row in df.iterrows():
            student = Student.objects.get(user__username=row['username'], school=school)
            subject, _ = Subject.objects.get_or_create(name=row['subject'], school=school)
            Score.objects.update_or_create(
                student=student,
                subject=subject,
                defaults={'ca_score': row['ca_score'], 'exam_score': row['exam_score']}
            )
        messages.success(request, "Scores uploaded successfully")
        return redirect('school_dashboard', school_id=school.id)
    return render(request, 'results/bulk_score_upload.html', {'school': school})



def generate_cumulative_report(request, student_id):
    student = Student.objects.get(id=student_id)
    scores = Score.objects.filter(student=student)
    
    # Aggregate by subject across all terms
    cumulative_scores = {}
    for s in scores:
        if s.subject.name not in cumulative_scores:
            cumulative_scores[s.subject.name] = {'ca':0, 'exam':0, 'count':0, 'teacher': s.subject.teacher.user.get_full_name()}
        cumulative_scores[s.subject.name]['ca'] += s.ca_score
        cumulative_scores[s.subject.name]['exam'] += s.exam_score
        cumulative_scores[s.subject.name]['count'] += 1

    # Compute average per subject
    final_scores = []
    total_marks = 0
    for subject, vals in cumulative_scores.items():
        ca_avg = vals['ca']/vals['count']
        exam_avg = vals['exam']/vals['count']
        total = ca_avg + exam_avg
        grade = 'A' if total>=70 else 'B' if total>=60 else 'C' if total>=50 else 'D' if total>=45 else 'F'
        total_marks += total
        final_scores.append({'subject': subject, 'ca': ca_avg, 'exam': exam_avg, 'total': total, 'grade': grade, 'teacher': vals['teacher']})

    average_score = total_marks / len(final_scores) if final_scores else 0

    # Use the same comment logic based on cumulative average
    import random
    if average_score >= 70: grade = 'A'
    elif average_score >= 60: grade = 'B'
    elif average_score >= 50: grade = 'C'
    elif average_score >= 45: grade = 'D'
    else: grade = 'F'

    principal_comments = {
        'A': [f"{student.user.first_name}, exceptional cumulative performance!"],
        'B': [f"{student.user.first_name}, good overall achievement."],
        'C': [f"{student.user.first_name}, satisfactory cumulative results."],
        'D': [f"{student.user.first_name}, needs improvement in several subjects."],
        'F': [f"{student.user.first_name}, poor overall performance."]
    }
    teacher_comments = {
        'A': [f"{student.user.first_name}, excellent class participation."],
        'B': [f"{student.user.first_name}, solid overall effort."],
        'C': [f"{student.user.first_name}, more effort required."],
        'D': [f"{student.user.first_name}, improvement needed."],
        'F': [f"{student.user.first_name}, must work harder."]
    }

    context = {
        'student': student,
        'scores': final_scores,
        'total_marks': total_marks,
        'average_score': round(average_score,2),
        'principal_comment': random.choice(principal_comments[grade]),
        'teacher_comment': random.choice(teacher_comments[grade])
    }

    from .utils import render_to_pdf
    pdf = render_to_pdf('results/termly_report.html', context)
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = f"{student.user.get_full_name()}_Cumulative_Report.pdf"
        response['Content-Disposition'] = f"inline; filename={filename}"
        return response
    return HttpResponse("Error generating cumulative PDF")



from django.contrib.auth.decorators import login_required
from accounts.models import Teacher, School
from results.models import Student


@login_required
def teacher_dashboard(request):
    teacher = Teacher.objects.get(user=request.user)
    school = teacher.school
    students = Student.objects.filter(school=school)
    return render(request, 'results/teacher_dashboard.html', {
        'teacher': teacher,
        'school': school,
        'students': students
    })


from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from .utils import render_to_pdf
from results.models import Student, Score, Psychomotor, Affective
import random
import qrcode
import base64
from io import BytesIO


def student_report_pdf(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    scores = Score.objects.filter(student=student)
    psycho = Psychomotor.objects.get(student=student)
    affective = Affective.objects.get(student=student)

    total_marks = sum([s.total for s in scores])
    average_score = total_marks / scores.count() if scores.exists() else 0

    best_subject = max(scores, key=lambda x: x.total).subject.name if scores.exists() else None
    least_subject = min(scores, key=lambda x: x.total).subject.name if scores.exists() else None

    # Grade for comment selection
    if average_score >= 70: grade = 'A'
    elif average_score >= 60: grade = 'B'
    elif average_score >= 50: grade = 'C'
    elif average_score >= 45: grade = 'D'
    else: grade = 'F'

    # Randomized comments
    principal_comments = {
        'A': [f"{student.user.first_name}, excellent performance! Keep it up!"],
        'B': [f"{student.user.first_name}, good job! Aim higher next term."],
        'C': [f"{student.user.first_name}, satisfactory performance, more effort is needed."],
        'D': [f"{student.user.first_name}, needs improvement. Let's work harder!"],
        'F': [f"{student.user.first_name}, poor performance. Immediate improvement required!"]
    }
    teacher_comments = {
        'A': [f"{student.user.first_name}, excellent class participation and results."],
        'B': [f"{student.user.first_name}, good effort, can still improve."],
        'C': [f"{student.user.first_name}, needs to put in more effort."],
        'D': [f"{student.user.first_name}, slow progress, extra support needed."],
        'F': [f"{student.user.first_name}, poor engagement. Immediate attention required."]
    }

    # Generate QR code
    qr = qrcode.QRCode(box_size=2, border=2)
    qr.add_data(f"Student ID: {student.id} - Total: {total_marks}")
    qr.make(fit=True)
    img = qr.make_image()
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    qr_code_base64 = base64.b64encode(buffer.getvalue()).decode()

    context = {
        'student': student,
        'scores': scores,
        'total_marks': total_marks,
        'average_score': round(average_score,2),
        'best_subject': best_subject,
        'least_subject': least_subject,
        'psycho': psycho,
        'affective': affective,
        'principal_comment': random.choice(principal_comments[grade]),
        'teacher_comment': random.choice(teacher_comments[grade]),
        'qr_code': qr_code_base64
    }

    pdf = render_to_pdf('results/student_report.html', context)
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = f"{student.user.get_full_name()}_Report.pdf"
        response['Content-Disposition'] = f"inline; filename={filename}"
        return response
    return HttpResponse("Error generating PDF")



import csv
from django.http import HttpResponse



def download_students_template(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=students_template.csv'
    writer = csv.writer(response)
    writer.writerow(['username','first_name','last_name','gender','student_class','email'])
    return response

@portal_required("results")
def download_scores_template(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=scores_template.csv'
    writer = csv.writer(response)
    writer.writerow(['username','subject','ca_score','exam_score','term'])
    return response


from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.core.paginator import Paginator
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from django.apps import apps
import datetime
import io
from .utils import SESSION_LIST
# PDF helper
from xhtml2pdf import pisa

# Excel
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Models (get dynamically - adjust app_label if models are elsewhere)
Score = apps.get_model('results', 'Score') if apps.is_installed('results') else None
Psychomotor = apps.get_model('results', 'Psychomotor')
Affective = apps.get_model('results', 'Affective')
Student = apps.get_model('students', 'Student') if apps.is_installed('students') else apps.get_model('results', 'Student')
School = apps.get_model('accounts', 'School')
SchoolClass = apps.get_model('students', 'SchoolClass')
Teacher = apps.get_model('staff', 'Teacher') if apps.is_installed('staff') else None
Subject = None
try:
    Subject = apps.get_model('results', 'Subject')
except Exception:
    Subject = None

# Helpers
def render_to_pdf(template_src, context_dict={}):
    html = render_to_string(template_src, context_dict)
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None


# Define term choices


@login_required
def results_dashboard(request, school_id):
    user = request.user
    school = get_object_or_404(School, id=school_id)

    teacher = getattr(user, 'teacher_profile', None)
    is_admin = user.is_staff or user.is_superuser

    # Filters
    class_id = request.GET.get('class_id') or ''
    subject_id = request.GET.get('subject_id') or ''
    term = request.GET.get('term') or ''
    session = request.GET.get('session') or ''
    student_q = request.GET.get('student') or ''

    # Classes
    if teacher and not is_admin:
        classes_qs = SchoolClass.objects.filter(teachers=teacher, school=school)
    else:
        classes_qs = SchoolClass.objects.filter(school=school)

    # Students
    students = Student.objects.filter(user__is_active=True, school_class__school=school)
    if not is_admin and teacher:
        teacher_class_ids = [c.id for c in classes_qs]
        students = students.filter(school_class__id__in=teacher_class_ids)
    if class_id:
        students = students.filter(school_class_id=class_id)
    if student_q:
        students = (
            students.filter(user__first_name__icontains=student_q) |
            students.filter(user__last_name__icontains=student_q) |
            students.filter(admission_no__icontains=student_q)
        )

    students = students.order_by('user__first_name', 'user__last_name')

    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(students, 20)
    try:
        students_page = paginator.page(page)
    except:
        students_page = paginator.page(1)

    student_ids = [s.id for s in students_page.object_list]

    # Subjects assigned to teacher via ClassSubjectTeacher
    if teacher and not is_admin:
        cst_qs = ClassSubjectTeacher.objects.filter(
            teacher=teacher,
            school_class__in=students_page.object_list.values_list('school_class', flat=True)
        )
        subjects_qs = Subject.objects.filter(
            id__in=cst_qs.values_list('subject_id', flat=True)
        )
    else:
        subjects_qs = Subject.objects.filter(school=school)

    # Scores filtered by teacher's subjects
    score_filters = {'student__id__in': student_ids}
    if subject_id:
        score_filters['subject_id'] = subject_id
    if term:
        score_filters['term'] = term
    if session:
        score_filters['session'] = session

    if teacher and not is_admin:
        score_qs = Score.objects.filter(**score_filters, subject__in=subjects_qs)
    else:
        score_qs = Score.objects.filter(**score_filters)

    # Psychomotor / Affective
    psy_qs = Psychomotor.objects.filter(student__id__in=student_ids)
    aff_qs = Affective.objects.filter(student__id__in=student_ids)

    # Lookup maps
    scores_map = {}
    for sc in score_qs:
        scores_map.setdefault(sc.student_id, []).append(sc)
    psy_map = {p.student_id: p for p in psy_qs}
    aff_map = {a.student_id: a for a in aff_qs}

    # Subject → teacher map
    subj_teacher_map = {}
    for cst in ClassSubjectTeacher.objects.filter(
        subject__in=subjects_qs, school_class__in=students_page.object_list.values_list('school_class', flat=True)
    ):
        subj_teacher_map[cst.subject_id] = cst.teacher.user.get_full_name() if cst.teacher else ''

    # Build rows
    rows = []
    for student in students_page.object_list:
        score_list = scores_map.get(student.id, [])
        totals = []
        subject_rows = []
        for sc in score_list:
            ca = getattr(sc, 'ca', 0) or 0
            exam = getattr(sc, 'exam', 0) or 0
            total = ca + exam
            totals.append(total)
            subject_name = getattr(sc, 'subject', None)
            teacher_name = subj_teacher_map.get(subject_name.id, '') if subject_name else ''
            subject_rows.append({
                'subject': str(subject_name) if subject_name else '',
                'ca': ca,
                'exam': exam,
                'total': total,
                'teacher': teacher_name
            })
        avg_total = sum(totals)/len(totals) if totals else None

        # Psychomotor / Affective
        psy = psy_map.get(student.id)
        aff = aff_map.get(student.id)
        psy_summary = [{'field': f, 'value': getattr(psy, f, None)} for f in ['neatness','agility','creativity','sports','handwriting']] if psy else []
        aff_summary = [{'field': f, 'value': getattr(aff, f, None)} for f in ['punctuality','cooperation','behavior','attentiveness','perseverance']] if aff else []

        rows.append({
            'student': student,
            'subject_rows': subject_rows,
            'avg_total': avg_total,
            'psychomotor': psy_summary,
            'affective': aff_summary
        })

    context = {
        'school': school,
        'classes': classes_qs,
        'subjects': subjects_qs,
        'students_page': students_page,
        'rows': rows,
        'paginator': paginator,
        'page_obj': students_page,
        'selected_class': int(class_id) if class_id else None,
        'selected_subject': int(subject_id) if subject_id else None,
        'selected_term': term,
        'selected_session': session,
        'student_q': student_q,
        'TERM_CHOICES': Score.TERM_CHOICES,
        'SESSION_LIST': SESSION_LIST,
    }

    return render(request, 'results/dashboard.html', context)




@login_required
def export_results_excel(request, school_id):
    """
    Export currently filtered dashboard to an Excel (xlsx) file.
    """
    school = get_object_or_404(School, id=school_id)
    # retrieve same filters as dashboard (simple approach)
    class_id = request.GET.get('class_id') or ''
    subject_id = request.GET.get('subject_id') or ''
    term = request.GET.get('term') or ''
    session = request.GET.get('session') or ''

    students = Student.objects.filter(school_class__school=school)
    if class_id:
        students = students.filter(school_class_id=class_id)
    students = students.order_by('user__first_name', 'user__last_name')

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = [
        "Admission No", "Student Name", "Class",
        "Subject", "CA", "Exam", "Total", "Grade",
        "Psychomotor(neatness,agility,creativity,sports,handwriting)",
        "Affective(punctuality,cooperation,behavior,attentiveness,perseverance)"
    ]
    ws.append(headers)

    # fetch required data in bulk
    student_ids = [s.id for s in students]
    score_qs = Score.objects.filter(student__id__in=student_ids) if Score else []
    psy_qs = Psychomotor.objects.filter(student__id__in=student_ids)
    aff_qs = Affective.objects.filter(student__id__in=student_ids)

    psy_map = {p.student_id: p for p in psy_qs}
    aff_map = {a.student_id: a for a in aff_qs}

    for student in students:
        row_base = [
            student.admission_no,
            student.full_name(),
            getattr(student.school_class, 'name', '')
        ]
        # For simplicity: join multiple subjects as separate lines or aggregated; here we list first subject if exists
        student_scores = score_qs.filter(student_id=student.id)
        if student_scores.exists():
            for sc in student_scores:
                ca = getattr(sc, 'ca', 0) or 0
                exam = getattr(sc, 'exam', 0) or 0
                total = ca + exam
                subject_name = str(getattr(sc, 'subject', '')) if Subject else ''
                psy = psy_map.get(student.id)
                aff = aff_map.get(student.id)
                psy_str = ",".join(str(getattr(psy, f)) for f in ['neatness','agility','creativity','sports','handwriting']) if psy else ''
                aff_str = ",".join(str(getattr(aff, f)) for f in ['punctuality','cooperation','behavior','attentiveness','perseverance']) if aff else ''
                row = row_base + [subject_name, ca, exam, total, '', psy_str, aff_str]
                ws.append(row)
        else:
            psy = psy_map.get(student.id)
            aff = aff_map.get(student.id)
            psy_str = ",".join(str(getattr(psy, f)) for f in ['neatness','agility','creativity','sports','handwriting']) if psy else ''
            aff_str = ",".join(str(getattr(aff, f)) for f in ['punctuality','cooperation','behavior','attentiveness','perseverance']) if aff else ''
            row = row_base + ['', '', '', '', '', psy_str, aff_str]
            ws.append(row)

    # Set some column widths
    for i, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    # prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    fname = f"{school.name}_results_{datetime.date.today().isoformat()}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response


# results/views.py
import io
import base64
import qrcode
from django.apps import apps
from django.conf import settings
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.utils import timezone

# dynamic models
Student = apps.get_model('students', 'Student') if apps.is_installed('students') else apps.get_model('results', 'Student')
Score = apps.get_model('results', 'Score') if apps.is_installed('results') else None
Psychomotor = apps.get_model('results', 'Psychomotor')
Affective = apps.get_model('results', 'Affective')
School = apps.get_model('accounts', 'School')
ResultVerification = apps.get_model('results', 'ResultVerification')

from .utils import grade_from_score


def _render_pdf_from_html(html):
    result = io.BytesIO()
    pdf = pisa.CreatePDF(io.BytesIO(html.encode('utf-8')), dest=result)
    if pdf.err:
        return None
    return result.getvalue()

def _render_pdf_response(template_src, context, filename):
    template = get_template(template_src)
    html = template.render(context)
    pdf_bytes = _render_pdf_from_html(html)
    if not pdf_bytes:
        return HttpResponse("Error generating PDF", status=500)
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def _get_verify_base(request):
    return getattr(settings, 'SITE_URL', request.build_absolute_uri('/').rstrip('/'))

def _get_school(student):
    # You told me student has direct relationship to school: student.school
    return getattr(student, 'school', None)

def report_card_view(request, student_id):
    """Render the report card (HTML). Also includes a Download PDF button."""
    student = get_object_or_404(Student, id=student_id)
    school = _get_school(student)

    # scores: include subject and subject teacher if present
    scores_qs = Score.objects.filter(student=student).select_related('subject') if Score else []
    scores = []
    totals = []
    for sc in scores_qs:
        ca = getattr(sc, 'ca', 0) or 0
        exam = getattr(sc, 'exam', 0) or 0
        total = ca + exam
        totals.append(total)
        teacher_name = ''
        subj = getattr(sc, 'subject', None)
        if subj:
            teacher = getattr(subj, 'teacher', None)
            teacher_name = teacher.user.get_full_name() if teacher else ''
        scores.append({
            'subject': str(subj) if subj else '',
            'ca': ca, 'exam': exam, 'total': total, 'grade': grade_from_score(total),
            'teacher': teacher_name
        })

    avg_total = sum(totals) / len(totals) if totals else 0
    best_subject = max(scores, key=lambda x: x['total']) if scores else None
    least_subject = min(scores, key=lambda x: x['total']) if scores else None

    # psychomotor/affective
    psychomotor = Psychomotor.objects.filter(student=student).first()
    affective = Affective.objects.filter(student=student).first()

    # verification token & QR
    verification_obj, created = ResultVerification.objects.get_or_create(student=student)
    verify_url = f"{_get_verify_base(request)}/results/verify/{student.admission_no}/?token={verification_obj.verification_token}"
    qr_data_uri = _generate_qr_data_uri(verify_url, box_size=6)

    # auto comments based on average
    if avg_total >= 70:
        teacher_comment = "Excellent performance. Keep it up."
        principal_comment = "Commendable performance."
    elif avg_total >= 50:
        teacher_comment = "Good work, keep improving."
        principal_comment = "Satisfactory; encourage more effort."
    else:
        teacher_comment = "Needs more attention and support."
        principal_comment = "Not acceptable. Intervention required."

    context = {
        'student': student,
        'school': school,
        'scores': scores,
        'avg_total': avg_total,
        'best_subject': best_subject,
        'least_subject': least_subject,
        'psychomotor': psychomotor,
        'affective': affective,
        'qr_data_uri': qr_data_uri,
        'verify_url': verify_url,
        'principal_name': getattr(school, 'principal_name', 'Principal') if school else 'Principal',
        'principal_title': getattr(school, 'principal_title', 'Principal') if school else 'Principal',
        'date_issued': timezone.localdate(),
        'teacher_comment': teacher_comment,
        'principal_comment': principal_comment,
    }

    return render(request, 'results/report_card.html', context)


def report_card_download(request, student_id):
    """Generate the PDF immediately and return as download."""
    student = get_object_or_404(Student, id=student_id)
    school = _get_school(student)

    # reuse logic from report_card_view but render to PDF response
    # build same context as above
    # (to avoid duplication, call report_card_view logic or refactor into helper—here inline for clarity)
    scores_qs = Score.objects.filter(student=student).select_related('subject') if Score else []
    scores = []
    totals = []
    for sc in scores_qs:
        ca = getattr(sc, 'ca', 0) or 0
        exam = getattr(sc, 'exam', 0) or 0
        total = ca + exam
        totals.append(total)
        teacher_name = ''
        subj = getattr(sc, 'subject', None)
        if subj:
            teacher = getattr(subj, 'teacher', None)
            teacher_name = teacher.user.get_full_name() if teacher else ''
        scores.append({
            'subject': str(subj) if subj else '',
            'ca': ca, 'exam': exam, 'total': total, 'grade': grade_from_score(total),
            'teacher': teacher_name
        })
    avg_total = sum(totals) / len(totals) if totals else 0
    best_subject = max(scores, key=lambda x: x['total']) if scores else None
    least_subject = min(scores, key=lambda x: x['total']) if scores else None
    psychomotor = Psychomotor.objects.filter(student=student).first()
    affective = Affective.objects.filter(student=student).first()
    verification_obj, created = ResultVerification.objects.get_or_create(student=student)
    verify_url = f"{_get_verify_base(request)}/results/verify/{student.admission_no}/?token={verification_obj.verification_token}"
    qr_data_uri = _generate_qr_data_uri(verify_url, box_size=6)

    if avg_total >= 70:
        teacher_comment = "Excellent performance. Keep it up."
        principal_comment = "Commendable performance."
    elif avg_total >= 50:
        teacher_comment = "Good work, keep improving."
        principal_comment = "Satisfactory; encourage more effort."
    else:
        teacher_comment = "Needs more attention and support."
        principal_comment = "Not acceptable. Intervention required."

    context = {
        'student': student,
        'school': school,
        'scores': scores,
        'avg_total': avg_total,
        'best_subject': best_subject,
        'least_subject': least_subject,
        'psychomotor': psychomotor,
        'affective': affective,
        'qr_data_uri': qr_data_uri,
        'verify_url': verify_url,
        'principal_name': getattr(school, 'principal_name', 'Principal') if school else 'Principal',
        'principal_title': getattr(school, 'principal_title', 'Principal') if school else 'Principal',
        'date_issued': timezone.localdate(),
        'teacher_comment': teacher_comment,
        'principal_comment': principal_comment,
    }

    filename = f"{student.full_name().replace(' ', '_')}_report_{context['date_issued']}.pdf"
    return _render_pdf_response('results/report_card.html', context, filename)



import io
import base64
import qrcode
from django.shortcuts import get_object_or_404, render


# -------------------------------
# Helper functions
# -------------------------------

def _generate_qr_data_uri(url, box_size=6):
    """
    Generate a base64 data URI of a QR code for the given URL.
    """
    qr = qrcode.QRCode(box_size=box_size, border=1)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode('ascii')


def _build_verification_url(student, token):
    """
    Construct the full verification URL for a student.
    """
    from django.conf import settings
    base_url = getattr(settings, "SITE_URL", "https://techcenter-p2au.onrender.com")
    return f"{base_url}/verify/{student.admission_no}/?token={token}"


# -------------------------------
# Main view
# -------------------------------

from django.shortcuts import get_object_or_404, render
from django.http import HttpResponseBadRequest
from django.conf import settings

def verify_result(request, admission_no):
    """
    Public verification of a student's result.

    URL format:
    /results/verify/<admission_no>/?token=<verification_token>&view=cumulative/term&term=<selected_term>
    """

    # ---------------------------
    # Fetch query parameters
    # ---------------------------
    token = request.GET.get("token", "").strip()
    view_type = request.GET.get("view", "term").lower()  # 'term' or 'cumulative'
    selected_term = request.GET.get("term")  # MUST be provided for term view

    student = get_object_or_404(Student, admission_no=admission_no)

    # ---------------------------
    # Verify token
    # ---------------------------
    verification = None
    if token:
        verification = ResultVerification.objects.filter(
            student=student,
            verification_token=token,
            valid=True
        ).first()
    is_verified = verification is not None

    # ---------------------------
    # Determine session
    # ---------------------------
    current_session = getattr(settings, "CURRENT_SESSION", None) \
        or getattr(student.school, "current_session", None) \
        or "2024/2025"

    # ---------------------------
    # Build result context
    # ---------------------------
    context = {}

    if view_type == "cumulative":
        # Cumulative result
        context = build_cumulative_result_context(student, session=current_session)
        context.update({
            "is_cumulative": True,
            "scores": [],  # cumulative uses subjects dict
            "status": "verified" if is_verified else "invalid",
            "token": token,
            "terms": context.get("terms", []),
            "show_ca": context.get("show_ca", True),
        })
    else:
        # Term view requires selected_term
        if not selected_term:
            return HttpResponseBadRequest("Term parameter is required for term view.")

        context = build_student_result_context(student, selected_term, session=current_session)
        context.update({
            "is_cumulative": False,
            "status": "verified" if is_verified else "invalid",
            "token": token,
            "subjects": {},  # not used in term view
            "terms": [selected_term],
            "show_ca": context.get("show_ca", True),
        })

    # ---------------------------
    # Ensure verification exists
    # ---------------------------
    if not verification:
        verification = ResultVerification.objects.create(student=student, valid=True)

    # ---------------------------
    # Generate QR code URL (safe)
    # ---------------------------
    base_url = getattr(settings, "SITE_URL", "https://techcenter-p2au.onrender.com")

    if context.get("is_cumulative"):
        # Cumulative view
        view_for_qr = "cumulative"
        verification_url = (
            f"{base_url}/results/verify/{student.admission_no}/"
            f"?token={verification.verification_token}"
            f"&view={view_for_qr}"
            f"&session={current_session}"
        )
    else:
        # Term view
        view_for_qr = "term"
        term_for_qr = context.get("terms")[0] if context.get("terms") else "1"  # safe fallback
        verification_url = (
            f"{base_url}/results/verify/{student.admission_no}/"
            f"?token={verification.verification_token}"
            f"&view={view_for_qr}"
            f"&term={term_for_qr}"
            f"&session={current_session}"
        )

    context["qr_data_uri"] = _generate_qr_data_uri(verification_url, box_size=6)

    # ---------------------------
    # Common template fields
    # ---------------------------
    school = student.school
    context.update({
        "principal_signature_url": getattr(school.principal_signature, 'url', None) if school else None,
        "student_photo_url": student.photo.url if student.photo else None,
        "school_logo_url": getattr(school.logo, 'url', None) if school else None,
        "selected_session": current_session,
    })

    return render(request, "results/verify_result.html", context)



   








# results/views.py (append these)
from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.apps import apps
from django.db.models import Q
from django.db.models import Prefetch
from .utils import SESSION_LIST

# dynamic model lookups (adapt to where your models live)
Student = apps.get_model('students', 'Student') if apps.is_installed('students') else apps.get_model('results', 'Student')
SchoolClass = apps.get_model('students', 'SchoolClass')
Score = None
try:
    Score = apps.get_model('results', 'Score')
except Exception:
    Score = None
Psychomotor = apps.get_model('results', 'Psychomotor')
Affective = apps.get_model('results', 'Affective')
Subject = None
try:
    Subject = apps.get_model('results', 'Subject')
except Exception:
    Subject = None




@login_required
def teacher_portal(request):
    """
    Teacher dashboard:
    - Shows classes assigned to teacher and their students
    - Tabs: Overview (students), Psych/Affective summary, Generate Reports
    - Filter by class, term, session, and search by student name/admission
    - Only show subjects assigned to this teacher
    """
    user = request.user
    teacher = getattr(user, 'teacher_profile', None)  # TeacherProfile instance

    if not teacher and not (user.is_staff or user.is_superuser):
        return render(request, 'results/teacher_forbidden.html', status=403)

    is_admin = user.is_staff or user.is_superuser

    # Classes the teacher can access
    if is_admin:
        classes_qs = SchoolClass.objects.all()
    else:
        try:
            classes_qs = teacher.classes.all()
        except Exception:
            classes_qs = SchoolClass.objects.filter(teacher=teacher)

    # Determine the school for bulk save
    school = None
    if not is_admin and teacher:
        first_class = classes_qs.first()
        if first_class:
            school = first_class.school
    elif is_admin:
        # For admin, pick first school (adjust if you want multi-school support)
        school = School.objects.first()

    # GET filters
    class_id = request.GET.get('class_id')
    term = request.GET.get('term')
    session = request.GET.get('session')
    q = request.GET.get('q', '').strip()

    TERM_LIST = [('1', 'Term 1'), ('2', 'Term 2'), ('3', 'Term 3')]

    # Base student queryset
    students = Student.objects.select_related('user', 'school_class', 'school')

    # Filter by class
    if class_id:
        students = students.filter(school_class_id=class_id)
    else:
        if not is_admin and teacher:
            cls_ids = list(classes_qs.values_list('id', flat=True))
            students = students.filter(school_class_id__in=cls_ids)

    # Search by name/admission
    if q:
        students = students.filter(
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q) |
            Q(admission_no__icontains=q)
        )

    students = students.order_by('user__first_name', 'user__last_name')

    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(students, 20)
    try:
        students_page = paginator.page(page)
    except:
        students_page = paginator.page(1)

    student_ids = [s.id for s in students_page.object_list]

    # GET SUBJECTS ASSIGNED TO TEACHER
    teacher_subject_ids = None
    if teacher and not is_admin:
        class_ids = list(students_page.object_list.values_list('school_class_id', flat=True))
        cst_qs = ClassSubjectTeacher.objects.filter(
            teacher=teacher,
            school_class_id__in=class_ids
        )
        teacher_subject_ids = list(cst_qs.values_list('subject_id', flat=True))

    # SCORE FILTERS
    score_filters = {"student_id__in": student_ids}
    if term:
        score_filters["term"] = term
    if session:
        score_filters["session"] = session

    if teacher_subject_ids is not None:
        score_qs = Score.objects.filter(**score_filters, subject_id__in=teacher_subject_ids).select_related('subject')
    else:
        score_qs = Score.objects.filter(**score_filters).select_related('subject')

    # PSYCHOMOTOR / AFFECTIVE
    psy_qs = Psychomotor.objects.filter(student_id__in=student_ids)
    aff_qs = Affective.objects.filter(student_id__in=student_ids)

    psy_map = {p.student_id: p for p in psy_qs}
    aff_map = {a.student_id: a for a in aff_qs}

    # BUILD SCORE MAP
    score_map = {}
    for sc in score_qs:
        score_map.setdefault(sc.student.id, []).append(sc)

    # PREPARE ROWS
    rows = []
    for student in students_page.object_list:
        student_scores = []
        for sc in score_map.get(student.id, []):
            ca = getattr(sc, 'ca', 0) or 0
            exam = getattr(sc, 'exam', 0) or 0
            total = ca + exam
            subj = getattr(sc, 'subject', None)
            teacher_name = ''
            if subj:
                cst = ClassSubjectTeacher.objects.filter(
                    subject=subj,
                    school_class=student.school_class
                ).first()
                if cst and cst.teacher:
                    teacher_name = cst.teacher.user.get_full_name()
            student_scores.append({
                'id': sc.id,
                'subject': str(subj) if subj else '',
                'ca': ca,
                'exam': exam,
                'total': total,
                'teacher': teacher_name
            })

        # Convert Psychomotor & Affective to dict for editable template
        psy_dict = {f: getattr(psy_map.get(student.id), f, 0) for f in ['neatness','agility','handwriting','sports','creativity']} if psy_map.get(student.id) else None
        aff_dict = {f: getattr(aff_map.get(student.id), f, 0) for f in ['punctuality','cooperation','behavior','attentiveness','perseverance']} if aff_map.get(student.id) else None

        rows.append({
            'student': student,
            'scores': student_scores,
            'psychomotor': psy_dict,
            'affective': aff_dict
        })

    context = {
        'teacher': teacher,
        'is_admin': is_admin,
        'classes': classes_qs,
        'students_page': students_page,
        'rows': rows,
        'paginator': paginator,
        'selected_class': int(class_id) if class_id else None,
        'term': term,
        'session': session,
        'TERM_LIST': TERM_LIST,
        'SESSION_LIST': SESSION_LIST,
        'q': q,
        'school': school,  # ✅ This fixes your bulk_save_all URL
    }

    return render(request, 'results/teacher_portal.html', context)







@login_required
def bulk_save_all(request, school_id):
    school = get_object_or_404(School, id=school_id)

    if request.method != "POST":
        messages.error(request, "Invalid request method.")
        return redirect('results:teacher_portal')

    # Capture session and term from POST
    session = request.POST.get("session")
    term = request.POST.get("term")

    if not session or not term:
        messages.error(request, "Missing session or term.")
        return redirect(request.META.get('HTTP_REFERER', 'results:teacher_portal'))

    # ----------------------
    # UPDATE SCORES
    # ----------------------
    for key, value in request.POST.items():
        if key.startswith("ca_") or key.startswith("exam_"):
            try:
                parts = key.split("_")
                field_type = parts[0]  # 'ca' or 'exam'
                score_id = parts[1]
                score = Score.objects.get(id=score_id, school=school)
                if field_type == "ca":
                    score.ca = float(value)
                    exam_value = request.POST.get(f"exam_{score_id}")
                    if exam_value:
                        score.exam = float(exam_value)
                elif field_type == "exam":
                    score.exam = float(value)
                    ca_value = request.POST.get(f"ca_{score_id}")
                    if ca_value:
                        score.ca = float(ca_value)
                # Recalculate total
                score.total = (score.ca or 0) + (score.exam or 0)
                score.save()
            except Score.DoesNotExist:
                continue
            except ValueError:
                messages.warning(request, f"Invalid number for score ID {score_id}.")
                continue

    # ----------------------
    # UPDATE PSYCHOMOTOR
    # ----------------------
    psy_fields = ["neatness", "agility", "handwriting", "sports", "creativity"]
    for student_id in request.POST.getlist("student_ids"):
        try:
            psycho = Psychomotor.objects.get(student_id=student_id, session=session, term=term, school=school)
        except Psychomotor.DoesNotExist:
            psycho = None

        if psycho:
            for field in psy_fields:
                form_field = f"psy_{field}_{student_id}"
                if form_field in request.POST:
                    try:
                        setattr(psycho, field, float(request.POST[form_field]))
                    except ValueError:
                        continue
            psycho.save()

    # ----------------------
    # UPDATE AFFECTIVE
    # ----------------------
    aff_fields = ["punctuality", "cooperation", "behavior", "attentiveness", "perseverance"]
    for student_id in request.POST.getlist("student_ids"):
        try:
            affect = Affective.objects.get(student_id=student_id, session=session, term=term, school=school)
        except Affective.DoesNotExist:
            affect = None

        if affect:
            for field in aff_fields:
                form_field = f"aff_{field}_{student_id}"
                if form_field in request.POST:
                    try:
                        setattr(affect, field, float(request.POST[form_field]))
                    except ValueError:
                        continue
            affect.save()

    messages.success(request, "Student results updated successfully!")
    return redirect(request.META.get('HTTP_REFERER', f'/results/teacher/portal/'))









@login_required
def teacher_student_detail(request, student_id):
    """
    Return a simple student details page (used by teacher portal 'View' action).
    You can use it to open a full report or edit psych/affective.
    """
    student = get_object_or_404(Student, id=student_id)
    psy = Psychomotor.objects.filter(student=student).first()
    aff = Affective.objects.filter(student=student).first()
    scores = Score.objects.filter(student=student).select_related('subject') if Score else []

    # build simple summary
    score_rows = []
    totals = []
    for sc in scores:
        ca = getattr(sc, 'ca', 0) or 0
        exam = getattr(sc, 'exam', 0) or 0
        total = ca + exam
        totals.append(total)
        score_rows.append({
            'subject': getattr(sc, 'subject', None),
            'ca': ca, 'exam': exam, 'total': total
        })
    avg = sum(totals)/len(totals) if totals else None

    return render(request, 'results/teacher_student_detail.html', {
        'student': student,
        'psychomotor': psy,
        'affective': aff,
        'scores': score_rows,
        'avg_total': avg,
    })



# results/views.py (append these imports at top of file if not present)
import io
import base64
import qrcode
from django.apps import apps
from django.conf import settings
from django.shortcuts import render, get_object_or_404, HttpResponse, redirect
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from openpyxl import Workbook
from django.utils import timezone
from django.contrib.auth.decorators import login_required

# utils inside this file (or import from your utils)
import io
import base64
import qrcode

def _generate_qr_data_uri(full_url, box_size=6):
    """
    Generate a base64 data URI of a QR code for a full URL.
    Always encodes the full URL so scanners open it directly.
    """
    qr = qrcode.QRCode(box_size=box_size, border=1)
    qr.add_data(full_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode('ascii')


def _pdf_from_html_string(html):
    result = io.BytesIO()
    pisa_status = pisa.CreatePDF(io.BytesIO(html.encode('utf-8')), dest=result)
    if pisa_status.err:
        return None
    return result.getvalue()

# Export helpers

@login_required
def _get_subject_and_students_for_teacher(request, subject_id):
    """
    Returns (subject, students_qs)
    The teacher must be owner of the subject (or fallback to subject class's students).
    """
    user = request.user
    teacher = getattr(user, 'teacher_profile', None)
    Subject = apps.get_model('results', 'Subject') if apps.is_installed('results') else None
    Student = apps.get_model('students', 'Student') if apps.is_installed('students') else apps.get_model('results','Student')

    if Subject:
        subject = get_object_or_404(Subject, id=subject_id)
        # try to ensure teacher has permission
        allowed = False
        if teacher:
            # common naming: subject.teacher or subject.teachers (M2M). handle both:
            subj_teacher = getattr(subject, 'teacher', None) or (getattr(subject, 'teachers', None) and subject.teachers.first())
            if subj_teacher and getattr(subj_teacher, 'id', None) == getattr(teacher, 'id', None):
                allowed = True
        # fallback: if user is staff/admin allow
        if user.is_staff or user.is_superuser:
            allowed = True

        if not allowed:
            # don't raise; return empty QS
            students_qs = Student.objects.none()
            return subject, students_qs

        # students for this subject — usually by class (subject assigned to a class)
        # Try to get subject.school_class or subject.classes, else select all students in teacher classes
        school_class = getattr(subject, 'school_class', None) or getattr(subject, 'classroom', None)
        if school_class:
            students_qs = Student.objects.filter(school_class=school_class).select_related('user','school')
        else:
            # final fallback: students where teacher == teacher (if you record that)
            students_qs = Student.objects.filter(school=teacher.school) if hasattr(teacher, 'school') else Student.objects.none()

        return subject, students_qs
    else:
        # no Subject model — fallback: return empty
        return None, Student.objects.none()


@login_required
def export_students_excel(request, subject_id):
    """
    Exports Excel for students assigned to the subject teacher.
    """
    Subject, Student = None, None
    subject, students_qs = _get_subject_and_students_for_teacher(request, subject_id)
    Student = apps.get_model('students', 'Student') if apps.is_installed('students') else apps.get_model('results','Student')
    Score = apps.get_model('results','Score') if apps.is_installed('results') else None
    Psychomotor = apps.get_model('results','Psychomotor')
    Affective = apps.get_model('results','Affective')

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    # header
    headers = ["Admission No", "Full Name", "Class", "Subject", "CA", "Exam", "Total", 
               "Neatness", "Agility", "Creativity", "Sports", "Handwriting",
               "Punctuality", "Cooperation", "Behavior", "Attentiveness", "Perseverance"]
    ws.append(headers)

    for student in students_qs:
        # gather scores for this student for this subject (if Score has subject FK)
        if Score and subject:
            scores = Score.objects.filter(student=student, subject=subject)
        elif Score:
            scores = Score.objects.filter(student=student)
        else:
            scores = []

        # get psych & aff
        psy = Psychomotor.objects.filter(student=student).first()
        aff = Affective.objects.filter(student=student).first()

        if scores:
            for sc in scores:
                ca = getattr(sc, 'ca', '') or ''
                exam = getattr(sc, 'exam', '') or ''
                total = (ca or 0) + (exam or 0) if isinstance(ca, (int,float)) or isinstance(exam, (int,float)) else ''
                row = [
                    getattr(student, 'admission_no', ''),
                    getattr(student, 'full_name', '') if callable(getattr(student,'full_name',None)) else getattr(student,'full_name','') ,
                    getattr(getattr(student,'school_class',None),'name',''),
                    getattr(getattr(sc,'subject',None),'name',''),
                    ca, exam, total,
                    getattr(psy,'neatness',''), getattr(psy,'agility',''), getattr(psy,'creativity',''), getattr(psy,'sports',''), getattr(psy,'handwriting',''),
                    getattr(aff,'punctuality',''), getattr(aff,'cooperation',''), getattr(aff,'behavior',''), getattr(aff,'attentiveness',''), getattr(aff,'perseverance','')
                ]
                ws.append(row)
        else:
            row = [
                getattr(student, 'admission_no', ''),
                getattr(student, 'full_name', '') if callable(getattr(student,'full_name',None)) else getattr(student,'full_name',''),
                getattr(getattr(student,'school_class',None),'name',''),
                getattr(subject,'name','') if subject else '',
                '', '', '',
                getattr(psy,'neatness',''), getattr(psy,'agility',''), getattr(psy,'creativity',''), getattr(psy,'sports',''), getattr(psy,'handwriting',''),
                getattr(aff,'punctuality',''), getattr(aff,'cooperation',''), getattr(aff,'behavior',''), getattr(aff,'attentiveness',''), getattr(aff,'perseverance','')
            ]
            ws.append(row)

    # prepare response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"students_subject_{subject_id}_{timezone.localdate()}.xlsx"
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



@login_required
def export_students_pdf(request, subject_id):
    """
    Simple list PDF (one page per student row) for teacher's students.
    """
    subject, students_qs = _get_subject_and_students_for_teacher(request, subject_id)
    Score = apps.get_model('results','Score') if apps.is_installed('results') else None
    Psychomotor = apps.get_model('results','Psychomotor')
    Affective = apps.get_model('results','Affective')

    rows = []
    for student in students_qs:
        scores = Score.objects.filter(student=student, subject=subject) if Score and subject else (Score.objects.filter(student=student) if Score else [])
        total = sum([(getattr(s,'ca',0) or 0) + (getattr(s,'exam',0) or 0) for s in scores]) if scores else 0
        rows.append({
            'student': student,
            'scores': scores,
            'total': total,
            'psychomotor': Psychomotor.objects.filter(student=student).first(),
            'affective': Affective.objects.filter(student=student).first(),
        })

    html = render_to_string('results/export_students_pdf.html', {
        'rows': rows,
        'subject': subject,
        'date': timezone.localdate(),
        'site_url': getattr(settings,'SITE_URL', request.build_absolute_uri('/').rstrip('/')),
    })
    pdf_bytes = _pdf_from_html_string(html)
    if not pdf_bytes:
        return HttpResponse("Error generating PDF", status=500)
    fname = f"students_subject_{subject_id}_{timezone.localdate()}.pdf"
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response

@login_required
def export_students_detailed_pdf(request, subject_id):
    """
    Detailed PDF: one full-report per student using your existing report_card template.
    Concatenates multiple student report HTMLs then renders to PDF.
    """
    subject, students_qs = _get_subject_and_students_for_teacher(request, subject_id)
    # reuse existing report_card.html to build concatenated HTML
    student_report_template = 'results/report_card.html'
    parts = []
    for student in students_qs:
        # ensure verification token exists & QR prepared (as earlier logic)
        ResultVerification = apps.get_model('results','ResultVerification')
        verification_obj, created = ResultVerification.objects.get_or_create(student=student)
        base = getattr(settings, 'SITE_URL', request.build_absolute_uri('/').rstrip('/'))
        verify_url = f"{base}/results/verify/{student.admission_no}/?token={verification_obj.verification_token}"
        qr_data_uri = _generate_qr_data_uri(verify_url, box_size=6)

        school = getattr(student, 'school', None)
        psychomotor = apps.get_model('results','Psychomotor').objects.filter(student=student).first()
        affective = apps.get_model('results','Affective').objects.filter(student=student).first()
        scores_qs = apps.get_model('results','Score').objects.filter(student=student) if apps.is_installed('results') else []

        ctx = {
            'student': student,
            'school': school,
            'scores': [{
                'subject': getattr(s,'subject',None).name if getattr(s,'subject',None) else '',
                'ca': getattr(s,'ca',0),
                'exam': getattr(s,'exam',0),
                'total': (getattr(s,'ca',0) or 0) + (getattr(s,'exam',0) or 0)
            } for s in scores_qs],
            'psychomotor': psychomotor,
            'affective': affective,
            'qr_data_uri': qr_data_uri,
            'verify_url': verify_url,
            'principal_name': getattr(school,'principal_name','Principal') if school else 'Principal',
            'principal_title': getattr(school,'principal_title','Principal') if school else 'Principal',
            'date_issued': timezone.localdate(),
            'teacher_comment': '',
            'principal_comment': '',
        }
        parts.append(render_to_string(student_report_template, ctx))

    # join parts (xhtml2pdf can render concatenated HTML)
    big_html = "<div style='page-break-after: always;'>".join(parts)
    pdf_bytes = _pdf_from_html_string(big_html)
    if not pdf_bytes:
        return HttpResponse("Error generating detailed PDF", status=500)
    fname = f"students_subject_detailed_{subject_id}_{timezone.localdate()}.pdf"
    resp = HttpResponse(pdf_bytes, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp



# results/views.py (append or merge)
import io
import base64
import qrcode
from django.apps import apps
from django.conf import settings
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, Http404
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Sum, F


ResultVerification = apps.get_model('results', 'ResultVerification')

# grading util (copy or import your existing)
def grade_from_score(total):
    if total >= 70: return 'A'
    if total >= 60: return 'B'
    if total >= 50: return 'C'
    if total >= 45: return 'D'
    return 'F'

def _generate_qr_data_uri(url, box_size=8):
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_Q,
        box_size=box_size,
        border=2,
    )
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()


def _pdf_from_html_string(html):
    result = io.BytesIO()
    pisa_status = pisa.CreatePDF(io.BytesIO(html.encode('utf-8')), dest=result)
    if pisa_status.err:
        return None
    return result.getvalue()




from django.db.models import Sum, F
from .utils import SESSION_LIST  
@portal_required("results")
@login_required
def student_portal(request):
    user = request.user
    student = getattr(user, 'student_profile', None)
    if not student:
        raise Http404("Student profile not found.")

    # ---------------- SESSION FILTERING (HARDCODED SOURCE) ----------------
    sessions = SESSION_LIST

    selected_session = request.GET.get("session")

    # Default to student's own session if nothing selected
    if not selected_session:
        selected_session = student.session

    # ---------------- SUBJECTS FILTERED BY SESSION ----------------
    subjects = Subject.objects.filter(
        scores__student=student,
        scores__session=selected_session
    ).distinct()

    term_list = [
        {"label": "First Term", "key": "first_term", "num": 1},
        {"label": "Second Term", "key": "second_term", "num": 2},
        {"label": "Third Term", "key": "third_term", "num": 3},
    ]

    terms_available = {}
    term_counts = {}
    completed_terms = 0

    cumulative_scores = {}

    # ---------------- BUILD CUMULATIVE SCORES ----------------
    for subject in subjects:
        cumulative_scores[subject.id] = {
            "subject": subject.name,
            "term_scores": {},
            "total": 0
        }
        for term in term_list:
            score_obj = Score.objects.filter(
                student=student,
                term=term['num'],
                subject=subject,
                session=selected_session
            ).first()

            total = (score_obj.ca or 0) + (score_obj.exam or 0) if score_obj else 0
            cumulative_scores[subject.id]["term_scores"][term['num']] = total
            cumulative_scores[subject.id]["total"] += total

    # ---------------- TERM COUNTS & AVAILABILITY ----------------
    for term in term_list:
        count_scores = Score.objects.filter(
            student=student,
            term=term['num'],
            session=selected_session
        ).count()

        term_counts[term['num']] = count_scores
        terms_available[term['key']] = count_scores > 0

        if count_scores > 0:
            completed_terms += 1

    progress_percent = int((completed_terms / 3) * 100)

    # ---------------- PREPARE RESULTS BUTTONS ----------------
    results = []
    for term in term_list:
        representative_score = Score.objects.filter(
            student=student,
            term=term['num'],
            session=selected_session
        ).first()

        results.append({
            'term': term['num'],
            'session': selected_session,
            'rep_score_id': representative_score.id if representative_score else None,
            'get_term_display': term['label']
        })

    # ---------------- BEST / WORST SUBJECTS ----------------
    best_subject_overall = max(cumulative_scores.values(), key=lambda x: x['total'], default=None)
    least_subject_overall = min(cumulative_scores.values(), key=lambda x: x['total'], default=None)

    overall_total = sum(x['total'] for x in cumulative_scores.values())
    avg_total = overall_total / len(subjects) if subjects else 0

    cumulative_available = any(terms_available.values())

    # ---------------- CONTEXT ----------------
    context = {
        'student': student,
        'school': student.school,
        'photo_url': student.photo.url if student.photo else None,

        'results': results,
        'sessions': sessions,
        'selected_session': selected_session,

        'term_list': term_list,
        'terms_available': terms_available,
        'term_counts': term_counts,
        'terms_completed': completed_terms,
        'progress_percent': progress_percent,
        'cumulative_available': cumulative_available,
        'cumulative_scores': cumulative_scores,
        'overall_total': overall_total,
        'avg_total': avg_total,
        'best_subject_overall': best_subject_overall,
        'least_subject_overall': least_subject_overall,
    }

    return render(request, 'results/student_portal.html', context)





import io
import base64
import random
import qrcode
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, Http404
from django.template.loader import get_template
from django.utils import timezone
from xhtml2pdf import pisa
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.db.models import Sum, F

# models (adjust import path if your apps differ)
from .models import Score, Psychomotor, Affective, ResultComment, ResultVerification
    # optional, only if you have separate app; else use student.school

# ---------- Grade function (uses your scale) ----------

# ---------- Randomized comments pool (5 each) ----------







import random
from .models import SchoolGradeComment
from .comments import PRINCIPAL_COMMENTS, TEACHER_COMMENTS
from .utils import get_pronouns
from itertools import cycle

# Store rotation state globally (memory-safe)
COMMENT_ROTATION_CACHE = {}


def rotate_list(key, items):
    """Rotate through a list without repeating the same item every time."""
    if key not in COMMENT_ROTATION_CACHE:
        COMMENT_ROTATION_CACHE[key] = cycle(items)
    return next(COMMENT_ROTATION_CACHE[key])


def apply_placeholders(template, student, term, session):
    pronouns = get_pronouns(student)

    return template.format(
        name=student.user.get_full_name(),
        first_name=student.user.first_name,
        term=str(term),
        session=str(session),
        pronoun=pronouns["pronoun"],
        possessive=pronouns["possessive"],
        objective=pronouns["objective"]
    )


def get_random_comment(school, student, grade, term, session, comment_type):
    """
    comment_type = 'principal' or 'teacher'
    Supports:
        - Rotation
        - Placeholders
        - School-saved comments first
        - Random default fallback
    """

    gs = GradeSetting.objects.filter(school=school).first()

    # ---------- 1️⃣ SCHOOL-SAVED COMMENTS ----------
    if gs:
        if comment_type == "principal":
            comment_list = (gs.principal_comments or {}).get(grade)
        else:
            comment_list = (gs.teacher_comments or {}).get(grade)

        if comment_list:
            if isinstance(comment_list, str):
                comment_list = [comment_list]  # convert to list

            # Rotate through list
            selected = rotate_list(f"{school.id}-{comment_type}-{grade}", comment_list)
            return apply_placeholders(selected, student, term, session)

    # ---------- 2️⃣ DEFAULT RANDOM FALLBACK ----------
    if comment_type == "principal":
        default_list = PRINCIPAL_COMMENTS.get(grade, [])
    else:
        default_list = TEACHER_COMMENTS.get(grade, [])

    if default_list:
        selected = random.choice(default_list)
        return apply_placeholders(selected, student, term, session)

    # ---------- 3️⃣ FINAL FALLBACK ----------
    return apply_placeholders("Keep improving, {name}.", student, term, session)




# results/views.py
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import GradeSettingForm
from .models import GradeSetting


@login_required
def grade_settings(request, school_id):
    if request.user.school.id != school_id:
        messages.error(request, "You cannot access another school's grade settings.")
        return redirect("school_admin:admin_dashboard")

    school = request.user.school
    grade_setting, _ = GradeSetting.objects.get_or_create(school=school)

    if request.method == "POST":
        form = GradeSettingForm(request.POST, instance=grade_setting)
        if form.is_valid():
            form.save()
            messages.success(request, "Grade settings updated successfully.")
            return redirect("results:grade_settings", school.id)
    else:
        form = GradeSettingForm(instance=grade_setting)

    context = {
        "form": form,
        "school": school,
    }
    return render(request, "results/grade_settings.html", context)


@login_required
def grade_settings_overview(request, school_id):
    # Ensure admin belongs to the school
    if request.user.school.id != school_id:
        messages.error(request, "Access denied.")
        return redirect("school_admin:admin_dashboard")

    school = request.user.school
    settings, _ = GradeSetting.objects.get_or_create(school=school)

    # Existing dictionaries
    grades = settings.grades or {}
    principal_comments = getattr(settings, "principal_comments", {})
    teacher_comments = getattr(settings, "teacher_comments", {})

    # Add grade interpretations from settings
    grade_interpretations = getattr(settings, "grade_interpretations", {})  # New field

    return render(request, "results/grade_settings_overview.html", {
        "school": school,
        "settings": settings,
        "grades": grades,
        "principal_comments": principal_comments,
        "teacher_comments": teacher_comments,
        "grade_interpretations": grade_interpretations,  # <-- Pass to template
    })



# ---------- helper: choose and save comments ----------
def get_grade_letter(avg):
    if avg >= 70: return 'A'
    if avg >= 60: return 'B'
    if avg >= 50: return 'C'
    if avg >= 45: return 'D'
    return 'F'

def _generate_qr_data_uri(url, box_size=8):
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_Q,
        box_size=box_size,
        border=2,
    )
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()


# ---------- MAIN VIEW (HTML) ----------
from django.db.models import F, Sum
from django.shortcuts import get_object_or_404
from django.utils import timezone
import random
from results.models import Score, Psychomotor, Affective, ResultComment, ResultVerification, ClassSubjectTeacher
from django.conf import settings

from django.db.models import F, Sum
from django.utils import timezone
from .models import (
    Score, Psychomotor, Affective, ResultComment, ResultVerification, ClassSubjectTeacher
)
from .utils import SESSION_LIST, interpret_grade  # import SESSION_LIST

def build_student_result_context(student, term, session):
    """
    Build the context for a student's result for a specific term and session only.
    NO fallback to another session.
    """
    if session not in SESSION_LIST:
        raise ValueError(f"Invalid session supplied: {session}")

    scores_qs = Score.objects.filter(
        student=student,
        term=term,
        session=session
    ).select_related("subject", "school")

    school = student.school
    scores = []
    overall_total = 0

    cst_qs = ClassSubjectTeacher.objects.filter(
        school_class=student.school_class,
        subject__in=[s.subject for s in scores_qs]
    ).select_related("teacher__user")
    cst_map = {cst.subject_id: cst for cst in cst_qs}

    class_has_ca = False  # flag to detect if any CA is set

    # Fetch school-specific grade interpretations
    grade_settings, _ = GradeSetting.objects.get_or_create(school=school)
    grade_interpretations_map = getattr(grade_settings, "grade_interpretations", {})

    # Compute class averages per subject
    subject_ids = [s.subject_id for s in scores_qs]
    class_avg_map = {}
    for sub_id in subject_ids:
        class_scores = Score.objects.filter(
            subject_id=sub_id,
            term=term,
            session=session,
            student__school_class=student.school_class
        )
        total_scores = [(cs.ca or 0) + (cs.exam or 0) for cs in class_scores]
        class_avg_map[sub_id] = sum(total_scores) / len(total_scores) if total_scores else 0

    for s in scores_qs:
        ca = s.ca or 0
        exam = s.exam or 0
        total = ca + exam
        overall_total += total

        if ca > 0:
            class_has_ca = True  # CA exists for at least one subject

        cst = cst_map.get(s.subject_id)
        teacher_name = cst.teacher.user.get_full_name() if cst and cst.teacher else "N/A"
        grade = grade_from_score_dynamic(total, school)

        # Add class average and school-specific grade interpretation
        class_average = class_avg_map.get(s.subject_id, 0)
        grade_interpretation = grade_interpretations_map.get(grade) or interpret_grade(grade)

        scores.append({
            "subject": s.subject.name,
            "ca": ca,
            "exam": exam,
            "total": total,
            "class_average": round(class_average, 2),
            "grade": grade,
            "grade_interpretation": grade_interpretation,
            "teacher": teacher_name,
        })

    avg = overall_total / len(scores) if scores else 0
    best_subject = max(scores, key=lambda x: x['total']) if scores else None
    least_subject = min(scores, key=lambda x: x['total']) if scores else None

    class_qs = Score.objects.filter(
        term=term,
        session=session,
        student__school_class=student.school_class
    ).values("student").annotate(total=Sum(F("ca") + F("exam")))

    ranking = sorted(class_qs, key=lambda x: x["total"] or 0, reverse=True)
    position = next((i for i, r in enumerate(ranking, start=1) if r["student"] == student.id), None)
    class_size = len(ranking)

    psychomotor = Psychomotor.objects.filter(
        student=student, term=term, session=session
    ).first()

    affective = Affective.objects.filter(
        student=student, term=term, session=session
    ).first()

    result_comment = ResultComment.objects.filter(
        student=student, term=term, session=session
    ).first()

    if not result_comment:
        final_grade = grade_from_score(avg)
        principal_comment = get_random_comment(
            student.school, student, final_grade, term, session, "principal"
        )
        teacher_comment = get_random_comment(
            student.school, student, final_grade, term, session, "teacher"
        )

        result_comment = ResultComment.objects.create(
            student=student,
            school=student.school,
            session=session,
            term=term,
            principal_comment=principal_comment,
            teacher_comment=teacher_comment
        )
    else:
        principal_comment = result_comment.principal_comment
        teacher_comment = result_comment.teacher_comment

    # Safely get a verification object
    verification_obj = (
        ResultVerification.objects
        .filter(student=student)
        .first()
    )
    if not verification_obj:
        verification_obj = ResultVerification.objects.create(student=student, valid=True)

    base = getattr(settings, "SITE_URL", "https://techcenter-p2au.onrender.com")
    verify_url = f"{base}/results/verify/{student.admission_no}/?token={verification_obj.verification_token}"
    qr_data_uri = _generate_qr_data_uri(verify_url, box_size=6)


    principal_signature_url = (
        school.principal_signature.url
        if school and school.principal_signature
        else None
    )
    

    school_logo_url = (
        school.logo.url
        if school and school.logo
        else None
    )


    return {
        "student": student,
        "school": school,
        "scores": scores,
        "overall_total": overall_total,
        "avg": avg,
        "position": position,
        "class_size": class_size,
        "best_subject": best_subject,
        "least_subject": least_subject,
        "psychomotor": psychomotor,
        "affective": affective,
        "principal_comment": principal_comment,
        "teacher_comment": teacher_comment,
        "qr_data_uri": qr_data_uri,
        "principal_signature_url": principal_signature_url,
        "school_logo_url": school_logo_url,
        "date_issued": timezone.localdate(),
        "sessions": SESSION_LIST,
        "selected_session": session,
        "selected_term": term,
        "class_has_ca": class_has_ca,  # <- template logic
    }




@login_required
def student_view_result(request, result_id):
    user = request.user
    student = getattr(user, "student_profile", None)
    if not student:
        raise Http404("Student profile not found.")

    if getattr(student, "is_result_blocked", False):
        return render(request, "results/result_blocked.html", {
            "student": student,
            "reason": student.block_reason or "Your result access has been restricted."
        })

    rep = get_object_or_404(Score.objects.select_related("student"), id=result_id)
    if rep.student_id != student.id:
        raise Http404("Not allowed.")

    selected_session = request.GET.get("session") or rep.session
    if selected_session not in SESSION_LIST:
        selected_session = rep.session
    selected_term = rep.term

    context = build_student_result_context(student, term=selected_term, session=selected_session)
    context.update({
        "term": selected_term,
        "session": selected_session,
        "sessions": SESSION_LIST,
    })

    return render(request, "results/student_result.html", context)




import io
import base64
from django.http import HttpResponse, Http404
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.shortcuts import get_object_or_404


@portal_required("results")
@login_required
def student_result_download(request, result_id):
    user = request.user
    student = getattr(user, "student_profile", None)

    if not student:
        raise Http404("Student profile not found.")

    # ---------- BACKEND BLOCK ENFORCEMENT ----------
    if getattr(student, "is_result_blocked", False):
        return render(request, "results/result_blocked.html", {
            "student": student,
            "reason": student.block_reason or "Your access to this result has been restricted."
        })
    # -----------------------------------------------

    # Get the specific score record
    rep = get_object_or_404(Score, id=result_id)
    if rep.student_id != student.id:
        raise Http404("Not allowed.")

    # ---------- SESSION HANDLING ----------
    selected_session = request.GET.get("session") or rep.session
    if selected_session not in SESSION_LIST:
        # fallback to rep.session if selected session is invalid
        selected_session = rep.session

    term = rep.term

    # Build context for PDF
    context = build_student_result_context(student, term, selected_session)
    context.update({
        "term": term,
        "selected_term": term,
        "session": selected_session,
        "selected_session": selected_session,
    })

    # ---------------- Add student photo as base64 ----------------
    student_photo_base64 = None
    try:
        # Adjust field name if your model uses another
        if student.passport:  # or student.photo, student.image, etc.
            with open(student.passport.path, "rb") as f:
                encoded = base64.b64encode(f.read()).decode()
                # detect extension for MIME type
                ext = student.passport.name.split(".")[-1].lower()
                mime = "jpeg" if ext in ["jpg","jpeg"] else ext
                student_photo_base64 = f"data:image/{mime};base64,{encoded}"
    except Exception:
        student_photo_base64 = None

    context["student_photo_base64"] = student_photo_base64

    # Render PDF
    template = get_template("results/student_result_pdf.html")
    html = template.render(context)

    pdf_file = io.BytesIO()
    pisa_status = pisa.CreatePDF(io.BytesIO(html.encode("utf-8")), dest=pdf_file)

    if pisa_status.err:
        return HttpResponse("Error generating PDF", status=500)

    pdf_file.seek(0)
    filename = f"{student.user.get_full_name().replace(' ', '_')}_result_{term}_{selected_session}.pdf"

    response = HttpResponse(pdf_file.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response





# imports at top of views.py (add if not present)
import io
import base64
import qrcode
from django.db.models import Sum, F
from django.template.loader import get_template
from django.http import HttpResponse, Http404
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from xhtml2pdf import pisa
from django.conf import settings
from django.utils import timezone


# -----------------------
# cumulative builder
# -----------------------
from collections import defaultdict
from django.db.models import Sum, F
from django.utils import timezone
from .models import Score, ResultComment, ResultVerification, Psychomotor, Affective, grade_from_score_dynamic
 


from collections import defaultdict
from django.db.models import F, Sum
from django.utils import timezone
from students.models import PromotionHistory



def build_cumulative_result_context(student, session=None):

    # ---------- USE SESSION_LIST ----------
    if session not in SESSION_LIST:
        session = SESSION_LIST[-1]

    terms = ["First", "Second", "Third"]
    TERM_MAP = {"First": "1", "Second": "2", "Third": "3"}

    subject_map = defaultdict(lambda: {
        "First": {"ca": 0, "exam": 0, "total": 0, "grade": "", "teacher": ""},
        "Second": {"ca": 0, "exam": 0, "total": 0, "grade": "", "teacher": ""},
        "Third": {"ca": 0, "exam": 0, "total": 0, "grade": "", "teacher": ""},
        "Average": 0,
        "AverageGrade": "",
    })

    overall_total = 0
    total_subject_count = 0
    psychomotor = {}
    affective = {}
    show_ca = False

    collected_comments = {
        "1": {"principal": "", "teacher": ""},
        "2": {"principal": "", "teacher": ""},
        "3": {"principal": "", "teacher": ""},
    }

    # ---------- COLLECT TERM RESULTS ----------
    for term in terms:
        term_code = TERM_MAP[term]
        term_context = build_student_result_context(student, term_code, session)

        for s in term_context.get("scores", []):
            subj_name = s["subject"]
            subject_map[subj_name][term] = {
                "ca": s["ca"],
                "exam": s["exam"],
                "total": s["total"],
                "grade": s["grade"],
                "teacher": s["teacher"],
            }

            if s["ca"] > 0:
                show_ca = True

        overall_total += sum(s["total"] for s in term_context.get("scores", []))
        total_subject_count += len(term_context.get("scores", []))

        if term_context.get("psychomotor"):
            pm = term_context["psychomotor"]
            psychomotor = {
                "neatness": pm.neatness,
                "agility": pm.agility,
                "creativity": pm.creativity,
                "sports": pm.sports,
                "handwriting": pm.handwriting,
            }

        if term_context.get("affective"):
            af = term_context["affective"]
            affective = {
                "punctuality": af.punctuality,
                "cooperation": af.cooperation,
                "behavior": af.behavior,
                "attentiveness": af.attentiveness,
                "perseverance": af.perseverance,
            }

        collected_comments[term_code]["principal"] = term_context.get("principal_comment") or ""
        collected_comments[term_code]["teacher"] = term_context.get("teacher_comment") or ""

    # ---------- COMPUTE SUBJECT AVERAGES ----------
    for subj, data in subject_map.items():
        totals = [data[t]["total"] for t in terms if data[t]["total"] > 0]
        avg = sum(totals) / len(totals) if totals else 0
        data["Average"] = round(avg, 2)
        data["AverageGrade"] = grade_from_score_dynamic(avg, student.school)

    avg_total = overall_total / total_subject_count if total_subject_count else 0
    overall_avg_grade = grade_from_score_dynamic(avg_total, student.school)

    # ---------- BEST / WEAK SUBJECT ----------
    subject_averages = {k: v["Average"] for k, v in subject_map.items() if v["Average"] > 0}
    best_subject = max(subject_averages, key=subject_averages.get) if subject_averages else None
    weak_subject = min(subject_averages, key=subject_averages.get) if subject_averages else None

    # ---------- CLASS POSITION ----------
    class_totals = (
        Score.objects
        .filter(student__school_class=student.school_class, session=session)
        .values("student")
        .annotate(total=Sum(F("ca") + F("exam")))
    )

    ranking = sorted(class_totals, key=lambda x: x["total"] or 0, reverse=True)
    position = next((i for i, r in enumerate(ranking, 1) if r["student"] == student.id), None)
    class_size = len(ranking)

    # ---------- COMMENTS ----------
    def pick_comment(field):
        return (
            collected_comments["3"][field]
            or collected_comments["2"][field]
            or collected_comments["1"][field]
        )

    principal_comment = pick_comment("principal")
    teacher_comment = pick_comment("teacher")

    # ---------- QR VERIFICATION ----------
    # Safely get a verification object
    verification_obj = ResultVerification.objects.filter(student=student).first()
    if not verification_obj:
         verification_obj = ResultVerification.objects.create(student=student, valid=True)

    # Build URL for cumulative QR code
    base = getattr(settings, "SITE_URL", "https://techcenter-p2au.onrender.com")
    verify_url = f"{base}/results/verify/{student.admission_no}/?token={verification_obj.verification_token}&cumulative=true"

    # Generate the QR code
    qr_data_uri = _generate_qr_data_uri(verify_url, box_size=6)


    # ---------- MEDIA ----------
    school = student.school
    principal_signature_url = school.principal_signature.url if school and school.principal_signature else None
    student_photo_url = student.photo.url if student.photo else None
    school_logo_url = school.logo.url if school and school.logo else None

    colspan_terms = 4 * len(terms)

    promotion_history = list(
        PromotionHistory.objects
        .filter(student=student)
        .order_by("-promoted_on")
    )

    # ---------- RETURN CONTEXT ----------
    return {
        "student": student,
        "school": school,
        "subjects": dict(subject_map),
        "terms": terms,
        "psychomotor": psychomotor,
        "affective": affective,
        "overall_total": overall_total,
        "avg_total": avg_total,
        "overall_avg_grade": overall_avg_grade,
        "best_subject": best_subject,
        "weak_subject": weak_subject,
        "position": position,
        "class_size": class_size,
        "principal_comment": principal_comment,
        "teacher_comment": teacher_comment,
        "qr_data_uri": qr_data_uri,
        "principal_signature_url": principal_signature_url,
        "student_photo_url": student_photo_url,
        "school_logo_url": school_logo_url,
        "date_issued": timezone.localdate(),
        "colspan_terms": colspan_terms,
        "selected_session": session,
        "show_ca": show_ca,
        "promotion_history": promotion_history,
        "is_cumulative": True,  # <-- useful for watermark/badge
    }





@login_required
def student_cumulative_result(request):
    user = request.user
    student = getattr(user, "student_profile", None)
    if not student:
        raise Http404("Student profile not found.")

    # ---------- BACKEND BLOCK ENFORCEMENT ----------
    if getattr(student, "is_result_blocked", False):
        return render(request, "results/result_blocked.html", {
            "student": student,
            "reason": student.block_reason or "Your cumulative result access has been restricted."
        })
    # -----------------------------------------------

    # ---------- SESSION HANDLING ----------
    system = SystemSetting.objects.first()
    current_session = system.current_session if system else None
    selected_session = request.GET.get("session") or current_session
    if selected_session not in SESSION_LIST:
        selected_session = current_session

    # ---------- BUILD CONTEXT ----------
    context = build_cumulative_result_context(student, selected_session)
    context.update({
        "session": selected_session,
        "terms": ['First', 'Second', 'Third'],
        "available_sessions": SESSION_LIST,  # for dropdown selection in template
    })

    return render(request, "results/student_cumulative_result.html", context)





from django.contrib.auth.decorators import login_required
from django.http import Http404, HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import io
from django.core.files.storage import default_storage
import base64
from django.shortcuts import render





def _get_image_base64(file_field):
    """Convert ImageField to base64 for xhtml2pdf."""
    if not file_field:
        return None
    try:
        if default_storage.exists(file_field.name):
            with default_storage.open(file_field.name, "rb") as f:
                data = f.read()
            return "data:image/png;base64," + base64.b64encode(data).decode()
        return None
    except Exception:
        return None

@portal_required("results")
@login_required
def student_cumulative_result_download(request):
    user = request.user
    student = getattr(user, "student_profile", None)
    if not student:
        raise Http404("Student profile not found.")

    # ---------- BACKEND BLOCK ENFORCEMENT ----------
    if getattr(student, "is_result_blocked", False):
        return render(request, "results/result_blocked.html", {
            "student": student,
            "reason": student.block_reason or "Your cumulative result access has been restricted."
        })
    # -----------------------------------------------

    # ---------- SESSION HANDLING ----------
    system = SystemSetting.objects.first()
    current_session = system.current_session if system else None
    selected_session = request.GET.get("session") or current_session
    if selected_session not in SESSION_LIST:
        selected_session = current_session

    # ---------- BUILD CONTEXT ----------
    context = build_cumulative_result_context(student, selected_session)

    # Convert images to base64 for PDF
    context.update({
        "school_logo_base64": _get_image_base64(student.school.logo) if student.school else None,
        "student_photo_base64": _get_image_base64(student.photo),
        "principal_signature_base64": _get_image_base64(student.school.principal_signature) if student.school else None,
        "session": selected_session,
    })

    # ---------- RENDER PDF ----------
    template = get_template("results/student_cumulative_result_pdf.html")
    html = template.render(context)

    pdf_file = io.BytesIO()
    pisa_status = pisa.CreatePDF(io.BytesIO(html.encode("utf-8")), dest=pdf_file, encoding='utf-8')
    if pisa_status.err:
        return HttpResponse("Error generating PDF", status=500)

    pdf_file.seek(0)
    filename = f"{student.user.get_full_name().replace(' ','_')}_cumulative_result_{selected_session}.pdf"

    response = HttpResponse(pdf_file.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
